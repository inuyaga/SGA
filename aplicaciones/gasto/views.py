from django.shortcuts import render
from django.views.generic import TemplateView, CreateView, UpdateView, ListView, DeleteView, DetailView
from aplicaciones.gasto.models import *
from aplicaciones.pago_proveedor.models import Pago
from aplicaciones.gasto.forms import *
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects

from django.urls import reverse_lazy
from django.shortcuts import redirect, reverse
from django.db import IntegrityError
from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.utils.formats import localize


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny

# CREACION DE EXCEL
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.styles import PatternFill
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# from decimal import Decimal

class InicioView(TemplateView):
    template_name = "gasto/inicio.html"


class CrearTipoGasto(PermissionRequiredMixin, CreateView):
    permission_required = 'gasto.add_tipogasto'
    model = TipoGasto
    template_name = "gasto/crear.html"
    form_class = TipoGastoForm
    success_url = reverse_lazy('gastos:tipoGastoList')


class UpdateTipoGasto(PermissionRequiredMixin, UpdateView):
    permission_required = 'gasto.change_tipogasto'
    model = TipoGasto
    template_name = "gasto/crear.html"
    form_class = TipoGastoForm
    success_url = reverse_lazy('gastos:tipoGastoList')


class TipoGastoList(PermissionRequiredMixin, ListView):
    permission_required = 'gasto.view_tipogasto'
    model = TipoGasto
    template_name = "gasto/list_tipo_gasto.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['TipoGastoForm'] = TipoGastoForm
        return context


class TipoGastoDelete(PermissionRequiredMixin, DeleteView):
    permission_required = 'gasto.delete_tipogasto'
    model = TipoGasto
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('gastos:tipoGastoList')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class GastoViewList(ListView):
    model = Gasto
    template_name = "gasto/list_gasto.html"
    paginate_by = 100

    def get_queryset(self):
        queryset = super().get_queryset()
        # queryset = queryset
        week = self.request.GET.get('week')
        departamento = self.request.GET.get('departamento')
        empresa = self.request.GET.get('empresa')
        tip_gasto = self.request.GET.get('tip_gasto')
        status = self.request.GET.get('status')
        Buscar = self.request.GET.get('Buscar')
        reembolsoID = self.request.GET.get('reembolsoID')

        if not self.request.user.is_superuser:
            permiso = self.request.user.has_perm('gasto.view_gasto')
            if permiso == False:
                try:
                    queryset = queryset.filter(g_depo=self.request.user.departamento)
                    print('try')
                except AttributeError as error:
                    messages.warning(self.request, 'Usuario:{} debe tener asignado un departamento para ver gastos de su departamento'.format(
                        self.request.user.username))

        if week != None and week != '':
            week = week.replace('W', '')
            week = week.split('-')
            queryset = queryset.filter(g_fechaCreacion__year=week[0])
            queryset = queryset.filter(g_fechaCreacion__week=week[1])                       

        if empresa != None and empresa != '':
            queryset = queryset.filter(g_depo__departamento_id_sucursal__sucursal_empresa_id__id=empresa)            
            
        if departamento != None and departamento != '':
            queryset = queryset.filter(g_depo=departamento)
        if tip_gasto != None and tip_gasto != '':
            queryset = queryset.filter(g_tipoGasto=tip_gasto)
        if status != '0' and status != None:            
            queryset = queryset.filter(g_estado=status)
        if Buscar != None and Buscar != '':
            queryset = queryset.filter(g_id=Buscar)
        if reembolsoID != None and reembolsoID != '':
            queryset = Reembolso.objects.get(r_id=reembolsoID).r_gastos.all()
            

        queryset=queryset.order_by('g_estado', '-g_id')
        return queryset

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        badge = {1: 'badge badge-info', 2: 'badge badge-success', 3: 'badge badge-primary', 4: 'badge badge-secondary', 5: 'badge badge-danger', 6: 'badge badge-warning',}
        context['form_filtro'] = FiltrosGastoForm(self.request.GET)        
        context['badge'] = badge

        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada

        return context

    def post(self, request, *args, **kwargs):
        checkId = request.POST.getlist('checkId')
        url = reverse('gastos:gasto_list')
        # print(checkId)
        if checkId:
            gastos = Gasto.objects.filter(g_id__in=checkId)
            reembolso_crear = Reembolso(r_by_user=request.user)
            reembolso_crear.save()
            reembolso_crear.r_gastos.add(*gastos)
            gastos.update(g_estado=6)
            messages.warning(request, 'Reembolso creado <a href="{}?reembolsoID={}">{}</a> '.format(
                url, reembolso_crear.pk, reembolso_crear.pk))
        else:
            messages.warning(request, 'No ha seleccionado ningun gasto..')

        return redirect(url)

class GastoRentaViewList(ListView):
    model = Pago
    template_name = "gasto/list_gasto_renta.html"
    paginate_by = 100

    def get_queryset(self):
        queryset = super().get_queryset()
        week = self.request.GET.get('week')
        status = self.request.GET.get('status')

        if not self.request.user.is_superuser:
            permiso = self.request.user.has_perm('gasto.view_gasto')
            if permiso == False:
                try:
                    queryset = queryset.filter(g_depo=self.request.user.departamento)
                    print('try')
                except AttributeError as error:
                    messages.warning(self.request, 'Usuario:{} debe tener asignado un departamento para ver gastos de su departamento'.format(
                        self.request.user.username))

        if week != None and week != '':
            week = week.replace('W', '')
            week = week.split('-')
            # queryset = queryset.filter(pago_creado__year=week[0])
            queryset = queryset.filter(pago_creado__week=week[1])

        if status != '0' and status != None:            
            queryset = queryset.filter(pago_status=status)

        queryset=queryset.order_by('pago_status', '-pago_id')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        badge = {1: 'badge badge-info', 2: 'badge badge-success', 3: 'badge badge-primary', 4: 'badge badge-secondary', 5: 'badge badge-danger', 6: 'badge badge-warning',}
        context['form_filtro'] = FiltroGastoRentaForm(self.request.GET)        
        context['badge'] = badge

        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada

        return context

    def post(self, request, *args, **kwargs):
        checkId = request.POST.getlist('checkId')
        url = reverse('gastos:gastorenta_list')
        # print(checkId)
        if checkId:
            gastos = Pago.objects.filter(pago_id__in=checkId)
            reembolso_crear = ReembolsoRenta(rr_by_user=request.user)
            reembolso_crear.save()
            reembolso_crear.rr_gastos.add(*gastos)
            gastos.update(pago_status=6)
            messages.warning(request, 'Reembolso creado con el folio <a href="{}?reembolsoID={}">{}</a> '.format(
                url, reembolso_crear.pk, reembolso_crear.pk))
        else:
            messages.warning(request, 'No ha seleccionado ningun gasto..')

        return redirect(url)


class GastoCreate(PermissionRequiredMixin, CreateView):
    permission_required = 'gasto.add_gasto'
    model = Gasto
    template_name = "gasto/crear.html"
    form_class = GastoForm
    success_url = reverse_lazy('gastos:gasto_list')


    def form_valid(self, form, **kwargs):
        context = self.get_context_data(**kwargs)
        
        try:
            self.request.user.departamento.departamento_id_sucursal.sucursal_empresa_id
            form.instance.g_depo = self.request.user.departamento
            form.instance.g_userCreador = self.request.user         

            instancia = form.save(commit=False)
            items_gastos = ItemGastoInlineFormSet(self.request.POST, self.request.FILES, instance=instancia)

            print(items_gastos.instance)
            if items_gastos.is_valid() and form.is_valid():
                instancia.save()
                items_gastos.save()
            else:
                return self.render_to_response(context)
        except AttributeError as identifier:
            
            messages.warning(self.request, 'Usuario:{} debe tener asignado un departamento para poder crear gastos error:{}'.format(
                self.request.user.username, identifier))
            return self.render_to_response(context)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ItemGastoInlineFormSet'] = ItemGastoInlineFormSet(self.request.POST or None, self.request.FILES or None)

        return context

    # def post(self, request, *args, **kwargs):
    #     context = self.get_context_data()
    #     context['ItemGastoInlineFormSet()']=ItemGastoInlineFormSet(request.POST)
    #     return super().post(request, *args, **kwargs)


class GastoUpdateView(PermissionRequiredMixin, UpdateView):
    permission_required = 'gasto.change_gasto'
    model = Gasto
    template_name = "gasto/crear.html"
    form_class = GastoForm
    success_url = reverse_lazy('gastos:gasto_list')

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)        
        context['ItemGastoInlineFormSet'] = ItemGastoInlineFormSet(self.request.POST or None,self.request.FILES or None, instance=self.object)
        return context
    
    def form_valid(self, form, **kwargs):    
        context = self.get_context_data(**kwargs)                    
        items_gastos = ItemGastoInlineFormSet(self.request.POST, self.request.FILES, instance=self.object)
        if items_gastos.is_valid():
            items_gastos.save()
        else:
            return self.render_to_response(context)

        return super().form_valid(form)


class GastoDetalleView(ListView):
    model = ItemGasto
    template_name = "gasto/detalle_gasto.html"
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(itm_gastID=self.kwargs.get('pk'))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['gasto']=Gasto.objects.get(g_id=self.kwargs.get('pk'))
        return context
    
    


class UpdateStatusView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = ()

    def post(self, request, *args, **kwargs):
        g_id = request.data["g_id"]
        g_estado = request.data["g_estado"]
        g_estado_filter = 1 if g_estado == 5 else g_estado - 1

        Gasto.objects.filter(g_id=g_id, g_estado=g_estado_filter).update(
            g_estado=g_estado)

        data = {
            'id': g_id,
            'new_stado': g_estado,
            'filter_stado': g_estado_filter,

        }
        return Response(data, status=status.HTTP_202_ACCEPTED)

class UpdateStatusRentaView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = ()

    def post(self, request, *args, **kwargs):
        g_id = request.data["g_id"]
        g_estado = request.data["g_estado"]
        g_estado_filter = 1 if g_estado == 5 else g_estado - 1

        Pago.objects.filter(pago_id=g_id, pago_status=g_estado_filter).update(
            pago_status=g_estado)

        data = {
            'id': g_id,
            'new_stado': g_estado,
            'filter_stado': g_estado_filter,
        }
        return Response(data, status=status.HTTP_202_ACCEPTED)


class GastoDelete(PermissionRequiredMixin, DeleteView):
    permission_required = 'gasto.delete_gasto'
    model = Gasto
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('gastos:gasto_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class ReembolsoList(PermissionRequiredMixin, ListView):
    permission_required = 'gasto.view_reembolso'
    model = Reembolso
    template_name = "gasto/reembolso.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class ReembolsoRentaList(PermissionRequiredMixin, ListView):
    permission_required = 'gasto.view_reembolso'
    model = ReembolsoRenta
    template_name = "gasto/reembolsoRenta.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class Download_report_reembolso(TemplateView):
    def get(self, request, *args, **kwargs):
        
        wb = Workbook()
        ws = wb.active

        IDReembolso = self.request.GET.get('IDReembolso')
        queryset = Reembolso.objects.all()

        if IDReembolso != None:
            queryset = queryset.get(r_id=IDReembolso)

        ws['A1'] = 'REEMBOLSO DE GASTOS #{}'.format(IDReembolso)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '###'        
        ws['B2'] = 'Creado'
        ws['C2'] = 'Estado'
        ws['D2'] = 'Empresa'
        ws['E2'] = 'Creador'
        ws['F2'] = 'Monto'        

        cont = 3

        for item in queryset.r_gastos.all():
            ws.cell(row=cont, column=1).value = item.g_id                    
            ws.cell(row=cont, column=1).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')                   
            ws.cell(row=cont, column=2).value = item.g_fechaCreacion
            ws.cell(row=cont, column=2).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
            ws.cell(row=cont, column=3).value = item.get_g_estado_display()
            ws.cell(row=cont, column=3).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
            ws.cell(row=cont, column=4).value = item.g_depo.departamento_id_sucursal.sucursal_empresa_id.empresa_nombre
            ws.cell(row=cont, column=4).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
            ws.cell(row=cont, column=5).value = item.g_userCreador.username
            ws.cell(row=cont, column=5).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
            ws.cell(row=cont, column=6).value = item.total_gasto()
            ws.cell(row=cont, column=6).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
            ws.cell(row=cont, column=6).number_format = '#,##0.00'
            cont += 1

            ws.cell(row=cont, column=2).value = "TIPO"
            ws.cell(row=cont, column=2).fill = PatternFill(start_color='4f92ff', end_color='4f92ff', fill_type='solid')
            ws.cell(row=cont, column=3).value = "MONTO"
            ws.cell(row=cont, column=3).fill = PatternFill(start_color='4f92ff', end_color='4f92ff', fill_type='solid')            
            ws.cell(row=cont, column=4).value = "FECHA"
            ws.cell(row=cont, column=4).fill = PatternFill(start_color='4f92ff', end_color='4f92ff', fill_type='solid')
            cont += 1
            for itemgasto in item.itemgasto_set.all():                
                ws.cell(row=cont, column=2).value = itemgasto.itm_tipoGasto.nombre
                ws.cell(row=cont, column=3).value = itemgasto.itm_monto                
                ws.cell(row=cont, column=4).value = itemgasto.itm_fecha
                cont += 1


        penultima_fila = cont-1
        ws.cell(row=cont, column=6).value = "Total"
        ws["F"+str(cont)] = "=SUM(F3:F{})".format(penultima_fila)
        ws["F"+str(cont)].number_format = '#,##0.00'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'reemboso.xlsx'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class Download_report_reembolsoRenta(TemplateView):
    def get(self, request, *args, **kwargs):
        
        wb = Workbook()
        ws = wb.active

        IDReembolso = self.request.GET.get('IDReembolso')
        queryset = ReembolsoRenta.objects.all()

        if IDReembolso != None:
            queryset = queryset.get(rr_id=IDReembolso)

        ws['A1'] = 'REEMBOLSO DE GASTOS DE RENTAS #{}'.format(IDReembolso)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '###'        
        ws['B2'] = 'Proveedor'
        ws['C2'] = 'Contrato'
        ws['D2'] = 'M Pago'
        ws['E2'] = 'Creado'
        ws['F2'] = 'Monto'        

        cont = 2
        suma=0
        ws.cell(row=cont, column=1).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        ws.cell(row=cont, column=2).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        ws.cell(row=cont, column=3).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        ws.cell(row=cont, column=4).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        ws.cell(row=cont, column=5).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        ws.cell(row=cont, column=6).fill = PatternFill(start_color='004ac2', end_color='004ac2', fill_type='solid')
        cont=cont+1
        for item in queryset.rr_gastos.all():
            ws.cell(row=cont, column=1).value = item.pago_id                    
            ws.cell(row=cont, column=2).value = item.contrato_id.contrato_proveedor_id.proveedor_nombre
            ws.cell(row=cont, column=3).value = item.contrato_id.contrato_sucursal.depto
            ws.cell(row=cont, column=4).value = item.pago_metodo
            ws.cell(row=cont, column=5).value = item.pago_creado
            ws.cell(row=cont, column=6).value = item.pago_monto
            ws.cell(row=cont, column=6).number_format = '#,##0.00'
            cont += 1
            suma += item.pago_monto or 0
            ws.cell(row=cont, column=6).value= suma
            


        # penultima_fila = cont-1
        ws.cell(row=cont, column=5).value = "Total"
        # ws["F"+str(cont)] = "=SUM(F3:F{})".format(penultima_fila)
        # ws["F"+str(cont)].number_format = '#,##0.00'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'reembosorenta.xlsx'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response




class Download_Gasto(TemplateView): 
    def get(self, request, *args, **kwargs):

        wb = Workbook()
        ws = wb.active
        
        queryset = Gasto.objects.all()
        
        week = self.request.GET.get('week')
        departamento = self.request.GET.get('departamento')
        empresa = self.request.GET.get('empresa')
        tip_gasto = self.request.GET.get('tip_gasto')
        status = self.request.GET.get('status')
        Buscar = self.request.GET.get('Buscar')
        reembolsoID = self.request.GET.get('reembolsoID')

        if not self.request.user.is_superuser:
            permiso = self.request.user.has_perm('gasto.view_gasto')
            if permiso == False:
                try:
                    queryset = queryset.filter(g_depo=self.request.user.departamento)
                    print('try')
                except AttributeError as error:
                    messages.warning(self.request, 'Usuario:{} debe tener asignado un departamento para ver gastos de su departamento'.format(
                        self.request.user.username))

        if week != None and week != '':
            week = week.replace('W', '')
            week = week.split('-')
            queryset = queryset.filter(g_fechaCreacion__year=week[0])
            queryset = queryset.filter(g_fechaCreacion__week=week[1])                       

        if empresa != None and empresa != '':
            queryset = queryset.filter(g_depo__departamento_id_sucursal__sucursal_empresa_id__id=empresa)            
            
        if departamento != None and departamento != '':
            queryset = queryset.filter(g_depo=departamento)
        if tip_gasto != None and tip_gasto != '':
            queryset = queryset.filter(g_tipoGasto=tip_gasto)
        if status != '0' and status != None:            
            queryset = queryset.filter(g_estado=status)
        if Buscar != None and Buscar != '':
            queryset = queryset.filter(g_id=Buscar)
        if reembolsoID != None and reembolsoID != '':
            queryset = Reembolso.objects.get(r_id=reembolsoID).r_gastos.all()

        queryset=queryset.order_by('g_estado', '-g_id')

        ws['A1'] = 'REPORTE DE GASTOS'
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '###'        
        ws['B2'] = 'Fecha'
        ws['C2'] = 'Sucursal'
        ws['D2'] = 'Usuario'
        ws['E2'] = 'Monto'
        ws['F2'] = 'Tipo gasto'        
        ws['G2'] = 'Estatus'        

        cont = 3

        for item in queryset:
            for detalle in item.itemgasto_set.all():
                ws.cell(row=cont, column=1).value = item.g_id                                        
                ws.cell(row=cont, column=2).value = localize(item.g_fechaCreacion)                                        
                ws.cell(row=cont, column=3).value = item.g_depo.departamento_id_sucursal.sucursal_nombre                                          
                ws.cell(row=cont, column=4).value = item.g_userCreador.username                                      
                ws.cell(row=cont, column=5).value = detalle.itm_monto                                          
                ws.cell(row=cont, column=6).value = detalle.itm_tipoGasto.nombre
                ws.cell(row=cont, column=7).value = item.get_g_estado_display()
                ws.cell(row=cont, column=5).number_format = '#,##0.00'
                cont += 1
            cont += 1



        penultima_fila = cont-1
        ws.cell(row=cont, column=4).value = "Total"
        ws["E"+str(cont)] = "=SUM(E3:E{})".format(penultima_fila)
        ws["E"+str(cont)].number_format = '#,##0.00'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'gastos.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
