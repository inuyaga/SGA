# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render, get_object_or_404
from aplicaciones.pago_proveedor.forms import NuevoProveedorForm, ContratosForms, PagoForms, CrearPagoForms, \
ContratosFormsEdit, NuevoDeptoCasaForm
from aplicaciones.pago_proveedor.models import Proveedor, Contrato, Pago, Renta
from django.shortcuts import redirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View, TemplateView
# from django.views.generic.list import ListView, CreateView, UpdateView
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
from django.urls import reverse_lazy
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects
from django.db.models import ProtectedError
from xml.dom.minidom import parse
from django.db import IntegrityError
from django.http import HttpResponse
# IMPORTACIONES PARA LOS REPORTES
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from django.http import FileResponse
from io import BytesIO
from reportlab.platypus import (SimpleDocTemplate, PageBreak, Image, Spacer,Paragraph, Table, TableStyle, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_RIGHT, TA_LEFT
from django.utils.formats import localize
PAGE_WIDTH = letter[0]
PAGE_HEIGHT = letter[1]
# Create your views here.
# def proveedor_report(render)
# pylint: disable = E1101

# def nuevo_proveedor(request):
#     if request.method == "POST":
#         form = NuevoProveedorForm(request.POST)
#         if form.is_valid():
#             # post = form.save(commit=False) Permite no guardar para agregar datos adicionales
#             # post.author = request.user
#             # post.published_date = timezone.now()
#             form.save()
#             return redirect('inicio:index')
#     else:
#         form = NuevoProveedorForm()
#         print(form)

#     contex = {
#         'usuario': request.user,
#         'proveedorForm': form,
#     }
#     return render(request, 'pagoproveedor/create_proveedor.html', contex)


# def lista_proveedores(request):
#     proveedor = Proveedor.objects.all()
#     contex = {
#         'proveedores': proveedor
#     }
#     return render(request, 'pagoproveedor/listar_proveedor.html', contex)


# def edicion_proveedor(request, id_proveedor):
#     prved = Proveedor.objects.get(id=id_proveedor)
#     if request.method == 'GET':
#         form = NuevoProveedorForm(instance=prved)
#     else:
#         form = NuevoProveedorForm(request.POST, instance=prved)
#         if form.is_valid():
#             form.save()
#             return redirect('proveedor:lista')
#     return render(request, 'pagoproveedor/create_proveedor.html', {'proveedorForm': form})

# CLASES PARA LA VISTA DE PROVEEDOR
# -------------------------------------------------------------------------------------------------------------
class ProveedorList(ListView):
    paginate_by = 20
    model = Proveedor
    template_name = 'pagoproveedor/listar_proveedor.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.view_proveedor',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProveedorList, self).dispatch(*args, **kwargs)

class ProveedorCreate(CreateView):
    model = Proveedor
    form_class = NuevoProveedorForm
    template_name = 'pagoproveedor/create_proveedor.html'
    success_url = reverse_lazy('proveedor:lista')

    def get_context_data(self, **kwargs):
        context = super(ProveedorCreate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Crear'
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.add_proveedor',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProveedorCreate, self).dispatch(*args, **kwargs)

class ProveedorUpdate(UpdateView):
    model = Proveedor
    form_class = NuevoProveedorForm
    template_name = 'pagoproveedor/create_proveedor.html'
    success_url = reverse_lazy('proveedor:lista')

    def get_context_data(self, **kwargs):
        context = super(ProveedorUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.change_proveedor',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProveedorUpdate, self).dispatch(*args, **kwargs)

class ProveedorDelete(DeleteView):
    model = Proveedor
    form_class = NuevoProveedorForm
    template_name = 'pagoproveedor/elimina_proveedor.html'
    success_url = reverse_lazy('proveedor:lista')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)

    @method_decorator(permission_required('pago_proveedor.delete_proveedor',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProveedorDelete, self).dispatch(*args, **kwargs)

# -------------------------------------------------------------------------------------------------------------


# CLASES PARA CONTRATOS
# -------------------------------------------------------------------------------------------------------------
class ContratoDetalle(DetailView):
    model=Contrato
    template_name = 'pagoproveedor/contrato/contrato_detalle.html'
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context



class ContratosList(ListView):
    paginate_by = 20
    model = Contrato
    template_name = 'pagoproveedor/contrato/contrato_list.html'

    @method_decorator(permission_required('pago_proveedor.view_contrato',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ContratosList, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context

    def get_queryset(self):
        queryset = super(ContratosList, self).get_queryset().order_by('-contrato_id')
        filter = self.request.GET.get('autorizado')
        filter2 = self.request.GET.get('descargar')
        if filter == 'True':
            queryset = queryset.filter(contrato_autorizado=True)
        elif filter== 'False':
            queryset = queryset.filter(contrato_autorizado=False)



        return queryset

class report_contratso(TemplateView):
    def get(self, request , *args, **kwargs):
        from openpyxl.styles import Font, Fill, Alignment
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws=wb.active

        ws['A1'] = 'Reporte de Contratos activos y autorizados'
        st=ws['A1']
        st.font = Font(size=14, b=True, color="FF0000")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'ID'
        ws['B2'] = 'Proveedor'
        ws['C2'] = 'Sucursal'
        ws['D2'] = 'Inicio'
        ws['E2'] = 'Termino'
        ws['F2'] = 'Monto'
        cont = 3
        contrato=Contrato.objects.filter(contrato_autorizado=True, contrato_status=True)
        for cto in contrato:
            ws.cell(row=cont, column=1).value = cto.contrato_id
            ws.cell(row=cont, column=2).value = str(cto.contrato_proveedor_id)
            ws.cell(row=cont, column=3).value = str(cto.contrato_sucursal)
            ws.cell(row=cont, column=4).value = cto.contrato_fecha_inicio
            ws.cell(row=cont, column=5).value = cto.contrato_fecha_termino
            ws.cell(row=cont, column=6).value = cto.contrato_monto
            ws.cell(row=cont, column=6).number_format = '#,##0'
            cont += 1

        ws["F"+str(cont)] = "=SUM(F3:F"+str(cont-1)+")"
        ws["F"+str(cont)].number_format = '#,##0'
        
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1



        nombre_archivo='reporte.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response




class ContratoCreate(CreateView):
    model = Contrato
    form_class = ContratosForms
    template_name = 'pagoproveedor/contrato/contrato_crear.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.add_contrato',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ContratoCreate, self).dispatch(*args, **kwargs)

class ContratoDelete(DeleteView):
    model = Contrato
    form_class = ContratosForms
    template_name = 'pagoproveedor/elimina_proveedor.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.delete_contrato',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ContratoDelete, self).dispatch(*args, **kwargs)

class ContratoUpdate(UpdateView):
    model = Contrato
    form_class = ContratosFormsEdit
    template_name = 'pagoproveedor/contrato/contrato_crear.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.change_contrato',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ContratoUpdate, self).dispatch(*args, **kwargs)
# -------------------------------------------------------------------------------------------------------------

# CLASES PARA LA VISTA DE PAGOS
# -------------------------------------------------------------------------------------------------------------
class PagoCreateP(CreateView):
    model = Pago
    form_class = PagoForms
    template_name = 'pagoproveedor/pago/pago_create.html'
    success_url = reverse_lazy('proveedor:contrato_listar')

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        try:
            context['pagos'] = Pago.objects.filter(contrato_id_id=self.kwargs.get('pk'))
        except Pago.DoesNotExist:
            context['pagos'] = None
        return context

    def form_valid(self, form):
        id_c = self.kwargs['pk']
        form.instance.contrato_id_id = id_c
        self.object = form.save()
        return super().form_valid(form)

class PagoCreate(View):
    id_pago=0
    pago_object=Contrato.objects.none()
    
    def myFirstPage(self, canvas, doc):
        Title = "SOLICITUD DE PAGO"
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)
        
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 50, Title)
        stylo = ParagraphStyle('firma_style', alignment=TA_CENTER, fontSize=6, fontName="Times-Roman")
        stylo2 = ParagraphStyle('firma_style', alignment=TA_CENTER, fontSize=8, fontName="Times-Bold")
        dta=[
            (Paragraph('', stylo2), Paragraph('RECIBIÓ', stylo2), Paragraph('', stylo2)),
            (Paragraph('', stylo), Paragraph('           ', stylo), Paragraph('', stylo)),
            ]

        tabla=Table(dta, colWidths=[6 * cm, 6 * cm, 6 * cm])
        tabla.wrapOn(canvas, PAGE_WIDTH, PAGE_HEIGHT)
        tabla.drawOn(canvas, 50, 20*cm)
        canvas.restoreState() 
    
    def dispatch(self, *args, **kwargs):
        self.id_pago=self.kwargs.get('pk')
        self.pago_object=Contrato.objects.get(contrato_id = self.id_pago)
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        new_asig = Pago(contrato_id_id=self.id_pago)
        new_asig.save()

        response = HttpResponse(content_type='application/pdf')
        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=18, title='SOLICITUD DE PAGO')
        items = []
        folio_format = ParagraphStyle('folio_serv', alignment = TA_LEFT, fontSize = 9, fontName="Times-Roman")
        fecha_format = ParagraphStyle('fecha_stylo', alignment = TA_RIGHT, fontSize = 9, fontName="Times-Roman")

        folio_txt="<strong>CONTRATO:</strong><em><u>{}</u></em>".format(self.id_pago)
        p1=Paragraph(folio_txt, folio_format)
        dta=[[p1]]

        folio_txt="<strong>SUCURSAL:</strong><em><u>{}</u></em>".format(localize(self.pago_object.contrato_sucursal))
        p2=Paragraph(folio_txt, fecha_format)
        dta=[(p1, p2)]

        tabla=Table(dta, colWidths=[9 * cm, 10 * cm])
        items.append(tabla)
        items.append(Spacer(0,20))
    
        dta=[
            ('PROVEEDOR:', self.pago_object.contrato_proveedor_id.proveedor_nombre),
            ('RFC:', self.pago_object.contrato_proveedor_id.proveedor_rfc),
            ('BANCO:', self.pago_object.contrato_proveedor_id.proveedor_banco),
            ('CUENTA:', self.pago_object.contrato_proveedor_id.proveedor_cuenta),
            ]

        tabla=Table(dta, colWidths=[5 * cm, 14 * cm])
        tabla.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, -1), 1, colors.dodgerblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.transparent)
            ]
        ))
        items.append(tabla)
        items.append(Spacer(0,20))

        doc.build(items, onFirstPage=self.myFirstPage) 
        response.write(buff.getvalue())
        buff.close()
        return response

class PagorRegerarPDF(View):
    id_contrato=0
    id_pago=0
    pago_object=Contrato.objects.none()
    new_asig=Pago.objects.none()
    
    def myFirstPage(self, canvas, doc):
        Title = "SOLICITUD DE PAGO"
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)
        
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 50, Title)
        stylo = ParagraphStyle('firma_style', alignment=TA_CENTER, fontSize=6, fontName="Times-Roman")
        stylo2 = ParagraphStyle('firma_style', alignment=TA_CENTER, fontSize=8, fontName="Times-Bold")
        dta=[
            (Paragraph('', stylo2), Paragraph('RECIBIÓ', stylo2), Paragraph('', stylo2)),
            (Paragraph('', stylo), Paragraph('           ', stylo), Paragraph('', stylo)),
            ]

        tabla=Table(dta, colWidths=[6 * cm, 6 * cm, 6 * cm])
        tabla.wrapOn(canvas, PAGE_WIDTH, PAGE_HEIGHT)
        tabla.drawOn(canvas, 50, 19.5*cm)
        canvas.restoreState() 
    
    def dispatch(self, *args, **kwargs):
        self.id_pago=self.kwargs.get('pago')
        self.new_asig = Pago.objects.get(pago_id=self.id_pago)
        self.pago_object=Contrato.objects.get(contrato_id = self.new_asig.contrato_id_id)
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):

        response = HttpResponse(content_type='application/pdf')
        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=18, title='SOLICITUD DE PAGO')
        items = []
        folio_format = ParagraphStyle('folio_serv', alignment = TA_LEFT, fontSize = 9, fontName="Times-Roman")
        fecha_format = ParagraphStyle('fecha_stylo', alignment = TA_RIGHT, fontSize = 9, fontName="Times-Roman")

        folio_txt="<strong>CONTRATO:</strong><em><u>{}</u></em>".format(self.pago_object.contrato_id)
        p1=Paragraph(folio_txt, folio_format)
        dta=[[p1]]

        folio_txt="<strong>SUCURSAL:</strong><em><u>{}</u></em>".format(localize(self.pago_object.contrato_sucursal))
        p2=Paragraph(folio_txt, fecha_format)
        dta=[(p1, p2)]

        tabla=Table(dta, colWidths=[9 * cm, 10 * cm])
        items.append(tabla)
        items.append(Spacer(0,20))
    
        dta=[
            ('PROVEEDOR:', self.pago_object.contrato_proveedor_id.proveedor_nombre),
            ('RFC:', self.pago_object.contrato_proveedor_id.proveedor_rfc),
            ('BANCO:', self.pago_object.contrato_proveedor_id.proveedor_banco),
            ('CUENTA:', self.pago_object.contrato_proveedor_id.proveedor_cuenta),
            ('MENSUALIDAD:', self.pago_object.contrato_monto),
            ('NOTA:', self.new_asig.pago_observacion),
            ]

        tabla=Table(dta, colWidths=[5 * cm, 14 * cm])
        tabla.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, -1), 1, colors.dodgerblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.transparent)
            ]
        ))
        items.append(tabla)
        items.append(Spacer(0,20))

        doc.build(items, onFirstPage=self.myFirstPage) 
        response.write(buff.getvalue())
        buff.close()
        return response

class PagoList(ListView):
    paginate_by = 10
    model = Pago
    # ordering = ['-factura_creado']
    template_name = 'pagoproveedor/pago/pago_list.html'
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context

class PagoDelete(DeleteView):
    model = Pago
    form_class = PagoForms
    template_name = 'pagoproveedor/pago/pago_delete.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    def get_success_url(self):
        pk=self.kwargs.get('id_pago')
        print(pk)
        url=reverse_lazy('proveedor:pago_crear', kwargs={'pk':pk})
        return url

    @method_decorator(permission_required('pago_proveedor.delete_pago',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PagoDelete, self).dispatch(*args, **kwargs)

class PagoUpdate(UpdateView):
    model = Pago
    form_class = PagoForms
    template_name = 'pagoproveedor/pago/pago_update.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        print(self.kwargs.get('pk'))
        try:
            context['pagos'] = Pago.objects.filter(contrato_id_id=self.kwargs.get('pk'))
            # context['pagos'].update(pago_status=3)
        except Pago.DoesNotExist:
            context['pagos'] = None
        return context

    @method_decorator(permission_required('pago_proveedor.change_pago',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PagoUpdate, self).dispatch(*args, **kwargs)

class PagoAutorizar(UpdateView):
    model = Pago
    form_class = PagoForms
    template_name = 'pagoproveedor/pago/pago_update.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        print(self.kwargs.get('pk'))
        try:
            context['pagos'] = Pago.objects.filter(contrato_id_id=self.kwargs.get('pk'))
            context['pagos'].update(pago_status=3)
            context['autorizar'] = "Verifica que la información sea correcta"
        except Pago.DoesNotExist:
            context['pagos'] = None
        return context

    @method_decorator(permission_required('pago_proveedor.pagoproveedor_autorizar_gasto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PagoAutorizar, self).dispatch(*args, **kwargs)

class PagoUpdateOb(UpdateView):
    model = Pago
    form_class = CrearPagoForms
    template_name = 'pagoproveedor/pago/pago_update.html'
    success_url = reverse_lazy('proveedor:contrato_listar')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        try:
            context['pagos'] = Pago.objects.filter(contrato_id_id=self.kwargs.get('pk'))
        except Pago.DoesNotExist:
            context['pagos'] = None
        return context

    @method_decorator(permission_required('pago_proveedor.change_pago',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PagoUpdateOb, self).dispatch(*args, **kwargs)



# -------------------------------------------------------------------------------------------------------------




# class ReportePersonasPDF(View):

# class GeneratePdf(View):
#     def get(self, request, *args, **kwargs):
#         contratos = Contrato.objects.all()
#         data = {
#              'contratos': contratos,
#              'today': '',
#              'amount': 39.99,
#             'customer_name': 'Cooper Mann',
#             'order_id': 1233434,
#         }
#         pdf = render_to_pdf('formatos/pdfejemplo.html', data)
#         return HttpResponse(pdf, content_type='application/pdf')


class CasaDeptoList(ListView):
    
    model = Renta
    ordering = ['depto']
    paginate_by = 25
    template_name = 'pagoproveedor/renta/lista_casadepto.html'
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        # context['object_list'] = Factura.objects.filter(factura_contrato_id__contrato_autorizado=True).select_related('factura_contrato_id')
        return context

    @method_decorator(permission_required('pago_proveedor.view_renta',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(CasaDeptoList, self).dispatch(*args, **kwargs)

class CasaDeptoAdd(CreateView):
    model = Renta
    form_class = NuevoDeptoCasaForm
    template_name = 'pagoproveedor/renta/add_casadepto.html'
    success_url = reverse_lazy('proveedor:depto_casa_lista')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.add_renta',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(CasaDeptoAdd, self).dispatch(*args, **kwargs)


class CasaDeptoUpdate(UpdateView):
    model = Renta
    form_class = NuevoDeptoCasaForm
    template_name = 'pagoproveedor/create_proveedor.html'
    success_url = reverse_lazy('proveedor:depto_casa_lista')

    def get_context_data(self, **kwargs):
        context = super(CasaDeptoUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pago_proveedor.change_renta',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(CasaDeptoUpdate, self).dispatch(*args, **kwargs)

class CasaDeptoDelete(DeleteView):
    model = Renta
    form_class = NuevoDeptoCasaForm
    template_name = 'pagoproveedor/elimina_proveedor.html'
    success_url = reverse_lazy('proveedor:depto_casa_lista')
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)

    @method_decorator(permission_required('pago_proveedor.delete_renta',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(CasaDeptoDelete, self).dispatch(*args, **kwargs)