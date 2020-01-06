from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView, View
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects
from aplicaciones.pedidos.models import (Area, Marca, Producto, Detalle_pedido, Pedido, Configuracion_pedido, Tipo_Pedido, Asignar_gasto_sucursal, Tipo_Pedido, 
Catalogo_Productos, Asignar_gasto_sucursal, STATUS, Inventario, CONTEO)
from aplicaciones.pedidos.froms import (AreaForm, MarcaForm, ProductoForm, PedidoForm, ProductoKitForm, ConfigForm, Tipo_PedidoForm, AsigGastoForm,
PedidoVentaForm, PedidoFacturaForm, PedidoSalidaForm, Catalogo_ProductosForm, InventarioResguardoForm, InventarioPikinForm,
InventarioOtrosForm, InventarioMermaForm, InventarioEdit) 
from aplicaciones.empresa.models import Pertenece_empresa
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Q
from django.db import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from aplicaciones.empresa.models import Sucursal

from django.db.models import ProtectedError

from django.http import JsonResponse

#Librerias reportlab a usar:
from django.http import HttpResponse
from io import BytesIO
from django.conf import settings
from reportlab.platypus import (SimpleDocTemplate, PageBreak, Image, Spacer,Paragraph, Table, TableStyle, Spacer, BaseDocTemplate, Frame, PageTemplate)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_RIGHT, TA_LEFT
PAGE_WIDTH = letter[0]
PAGE_HEIGHT = letter[1]

from aplicaciones.empresa.models import Departamento
# Create your views here.
# pylint: disable=no-member
# pylint: disable = E1101

# VISTAS GENERICAS DE LA AREA
# ---------------------------------------------------------------------------------------------------------
class AreaCreate(CreateView):
    model = Area
    form_class = AreaForm
    template_name = "pedidos/area/area_create.html"
    success_url = reverse_lazy('pedidos:listar_area')

    def get_context_data(self, **kwargs):
        context = super(AreaCreate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Crear'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.add_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(AreaCreate, self).dispatch(*args, **kwargs)

class AreaList(ListView):
    paginate_by = 20
    model = Area
    template_name='pedidos/area/area_list.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pedidos.view_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(AreaList, self).dispatch(*args, **kwargs)

class AreaUpdate(UpdateView):
    model = Area
    form_class = AreaForm
    template_name = "pedidos/area/area_create.html"
    success_url = reverse_lazy('pedidos:listar_area')

    def get_context_data(self, **kwargs):
        context = super(AreaUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.change_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(AreaUpdate, self).dispatch(*args, **kwargs)

class AreaDelete(DeleteView):
    model = Area
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:listar_area')

    def get_context_data(self, **kwargs):
        context = super(AreaDelete, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    @method_decorator(permission_required('pedidos.delete_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(AreaDelete, self).dispatch(*args, **kwargs)

# ---------------------------------------------------------------------------------------------------------

# VISTAS GENERICAS DE LA MARCAS
# ---------------------------------------------------------------------------------------------------------
class MarcaCreate(CreateView):
    model = Marca
    form_class = MarcaForm
    template_name = "pedidos/marca/marca_create.html"
    success_url = reverse_lazy('pedidos:listar_marca')

    def get_context_data(self, **kwargs):
        context = super(MarcaCreate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Crear'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.add_marca',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MarcaCreate, self).dispatch(*args, **kwargs)


class MarcaList(ListView):
    paginate_by = 20
    model = Marca
    template_name='pedidos/marca/marca_list.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        return context
    @method_decorator(permission_required('pedidos.view_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MarcaList, self).dispatch(*args, **kwargs)

class MarcaUpdate(UpdateView):
    model = Marca
    form_class = MarcaForm
    template_name = "pedidos/marca/marca_create.html"
    success_url = reverse_lazy('pedidos:listar_marca')

    def get_context_data(self, **kwargs):
        context = super(MarcaUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.change_marca',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MarcaUpdate, self).dispatch(*args, **kwargs)

class MarcaDelete(DeleteView):
    model = Marca
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:listar_marca')

    def get_context_data(self, **kwargs):
        context = super(MarcaDelete, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Eliminar'
        context['usuario'] = self.request.user

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    @method_decorator(permission_required('pedidos.delete_area',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MarcaDelete, self).dispatch(*args, **kwargs)
# ---------------------------------------------------------------------------------------------------------


# VISTAS GENERICAS DE LA PRODUCTOS
# ---------------------------------------------------------------------------------------------------------
class ProductoCreate(CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = "pedidos/producto/producto_create.html"
    success_url = reverse_lazy('pedidos:listar_producto')

    def get_context_data(self, **kwargs):
        context = super(ProductoCreate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Crear'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.add_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductoCreate, self).dispatch(*args, **kwargs)

class ProductokitCreate(CreateView):
    model = Producto
    form_class = ProductoKitForm
    template_name = "pedidos/producto/producto_create.html"
    success_url = reverse_lazy('pedidos:listar_producto')

    def get_context_data(self, **kwargs):
        context = super(ProductokitCreate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Crear'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.add_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductokitCreate, self).dispatch(*args, **kwargs)

class ProductoList(ListView):
    paginate_by = 20
    model = Producto
    template_name='pedidos/producto/list_producto.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        # context['tipo_ped_list'] = Tipo_Pedido.objects.all()

        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super(ProductoList, self).get_queryset()
        producto_categoria=self.request.GET.getlist('producto_categoria')
        tipo_producto=self.request.GET.getlist('tipo_producto')
        checkbox_visible=self.request.GET.getlist('checkbox_visible')
        Buscar=self.request.GET.get('Buscar')
        if len(tipo_producto) != 0:
            queryset=queryset.filter(tipo_producto__in=tipo_producto)
        if len(producto_categoria) != 0:
            queryset=queryset.filter(producto_categoria__in=producto_categoria)
        if len(checkbox_visible) != 0:
            queryset=queryset.filter(producto_visible__in=checkbox_visible)
        if Buscar != None:
            queryset=queryset.filter(Q(producto_codigo=Buscar) | Q(producto_nombre__icontains=Buscar))


        return queryset

    @method_decorator(permission_required('pedidos.view_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductoList, self).dispatch(*args, **kwargs)


class ProductoUpdate(UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = "pedidos/producto/producto_create.html"
    success_url = reverse_lazy('pedidos:listar_producto')

    def get_context_data(self, **kwargs):
        context = super(ProductoUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.change_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductoUpdate, self).dispatch(*args, **kwargs)

class ProductoKitUpdate(UpdateView):
    model = Producto
    form_class = ProductoKitForm
    template_name = "pedidos/producto/producto_create.html"
    success_url = reverse_lazy('pedidos:listar_producto')

    def get_context_data(self, **kwargs):
        context = super(ProductoKitUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedidos.change_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductoKitUpdate, self).dispatch(*args, **kwargs)

class ProductoDelete(DeleteView):
    model = Producto
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:listar_producto')

    def get_context_data(self, **kwargs):
        context = super(ProductoDelete, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Eliminar'
        context['usuario'] = self.request.user

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'mns_producto': 'Este producto esta relacionado con una compra'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)

    @method_decorator(permission_required('pedidos.delete_producto',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ProductoDelete, self).dispatch(*args, **kwargs)



class ProductoCompraList(DetailView):
    # paginate_by = 20
    context_object_name = 'objeto'
    model = Tipo_Pedido
    template_name='pedidos/compra_tiendas.html'

    def get_context_data(self, **kwargs):
        from datetime import datetime, date
        import calendar
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        tipo_pedido=self.kwargs.get('pk')
        context['msn_empresa'] = None
        context['conteo'] = Detalle_pedido.objects.filter(detallepedido_creado_por=self.request.user, detallepedido_status=False).count()

        # ESTABLECEMOS LA FECHA ACTUAL
        today = datetime.now()
        # CONSULTAMOS CUAL ES EL ULTIMO DIA DEL MES ACTUAL
        last_day=calendar.monthrange(today.year, today.month)[1]
        # INICIALIZAMOS LA FECHA INICIAL
        start_date = datetime(today.year, today.month, 1)
        # INICIALIZAMOS LA FECHA FINAL
        end_date = datetime(today.year, today.month, last_day)

        try:
            empresa_pertenece = Pertenece_empresa.objects.get(pertenece_id_usuario=self.request.user)
            conteo_pedido=Pedido.objects.filter(pedido_tipoPedido=tipo_pedido, pedido_id_depo=empresa_pertenece.pertenece_empresa, pedido_fecha_pedido__range=(start_date,end_date)).exclude(pedido_status=3).count()
            context['pedidos_del_mes'] = conteo_pedido
        except ObjectDoesNotExist as error:
            context['msn_empresa'] = 'Para poder hacer pedidos es nesesario que pertenesca a una empresa, ¡comuniquese con el administrador del sistema!'

        return context

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ProductoCompraList, self).dispatch(request, *args, **kwargs)


    def post(self, request, *args, **kwargs):
        from decimal import Decimal
        codigo = request.POST.get('nombre')
        cantidad = request.POST.get('catidad')
        tipo_mensaje=True
        mensaje=''
        tipo_pedido=self.kwargs.get('pk')
        print(tipo_pedido)


        if Decimal(cantidad) <= 0:
            mensaje='Ingrese un numero positivo por favor'
            tipo_mensaje=False
        else:
            # OBTENER ID DE TIPO PEDIDO

            get_tipo_pedido=Tipo_Pedido.objects.get(tp=tipo_pedido)
            # OBTENER QUE DEPARTAMENTO PERTENECE
            get_depo = Pertenece_empresa.objects.get(pertenece_id_usuario=request.user)
            # OBTENER CODIGO DE PRODUCTO EN ESPECIFICO
            producto = Producto.objects.get(producto_codigo=codigo)
            # LIMITE DE GASTO DEL DEPARTAMENTO
            try:
                get_gasto=Asignar_gasto_sucursal.objects.get(ags_tipo_ped=tipo_pedido,ags_sucursal=get_depo.pertenece_empresa)
                limite_gasto=get_gasto.ags_maximo_gasto
                # OBTENEMOS LA SUMA TOTAL DE PEDIDO EN EL MOMENTO PREVIO AL PEDIDO EN ESPECIFICO DETALLE DE PEDIDO
                suma_pedido=Detalle_pedido.objects.filter(detallepedido_creado_por=request.user, detallepedido_status=False, detallepedido_tipo_pedido=tipo_pedido).aggregate(suma_total=Sum(F('detallepedido_precio') * F('detallepedido_cantidad')))
                # SUMAMOS LA SUMA DEL PEDIDO CON EL PRODUCTO ACTUAL ÁRA AGREGAR
                subtotal=(Decimal(producto.producto_precio) * Decimal(cantidad)) + Decimal(0 if suma_pedido['suma_total'] == None else suma_pedido['suma_total'])

                if subtotal <= limite_gasto:
                    if producto.producto_kit:
                        queryset = producto.producto_productos.all()
                        for producto_list in queryset:
                            GuardaDetallePedido(
                                producto_list.producto_codigo,
                                cantidad,
                                request.user,
                                producto_list.producto_precio,
                                tipo_pedido)
                        mensaje='Agregado Correctamente'
                    else:
                        GuardaDetallePedido(
                            producto.producto_codigo,
                            cantidad,
                            request.user,
                            producto.producto_precio,
                            tipo_pedido)
                        mensaje='Agregado Correctamente'
                else:
                    tipo_mensaje=False
                    mensaje='Supera el limite admitido para '+ str(get_tipo_pedido.tp_nombre)
            except ObjectDoesNotExist as NoExiste:
                tipo_mensaje=False
                mensaje='Por el momento no ha sido asignado a un departamento o no se a asignado un gasto a la sucursal, contacte con el administrador'



        pedido_conteo=Detalle_pedido.objects.filter(detallepedido_creado_por=request.user, detallepedido_status=False).count()
        json = JsonResponse(
            {
                'content':{
                    'mensaje': mensaje,
                    'conteo': pedido_conteo,
                    'tipo_mensaje': tipo_mensaje,
                }
            }
        )

        return json

def GuardaDetallePedido(codigo, cantidad, creado_por, precio, categoria):
    detalle_pedido=Detalle_pedido(
    detallepedido_producto_id_id=codigo,
    detallepedido_cantidad=cantidad,
    detallepedido_creado_por=creado_por,
    detallepedido_precio=precio,
    detallepedido_tipo_pedido=categoria,
    )
    detalle_pedido.save()

def Obtenernumero(numero):
    if numero == None:
        return 0
    else:
        return numero

class DetalleList(ListView):
    paginate_by = 10
    model = Detalle_pedido
    template_name = 'pedidos/listado_pre_pedido.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['conteo'] = Detalle_pedido.objects.filter(detallepedido_creado_por=self.request.user, detallepedido_status=False).count()
        context['tipos_pe_list'] = Tipo_Pedido.objects.all()
        context['colors_butons'] = ['default', 'primary', 'success', 'info']

        return context
    def get_queryset(self):
        queryset = super(DetalleList, self).get_queryset()
        return queryset.filter(detallepedido_creado_por=self.request.user, detallepedido_status=False)

    @method_decorator(permission_required('pedidos.view_detalle_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(DetalleList, self).dispatch(*args, **kwargs)

class DetalleDelete(DeleteView):
    model = Detalle_pedido
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:pedido_tienda_listado')
    def get_context_data(self, **kwargs):
        context = super(DetalleDelete, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Eliminar'
        context['usuario'] = self.request.user

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    @method_decorator(permission_required('pedidos.delete_detalle_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(DetalleDelete, self).dispatch(*args, **kwargs)

class Crear_pedido_tiendaView(View):
    def get(self, request, *args, **kwargs):
        empresa_pertenece = Pertenece_empresa.objects.get(pertenece_id_usuario=request.user)
        departamento=empresa_pertenece.pertenece_empresa.departamento_id_depo
        tipo_pedido=self.kwargs.get('tipo')

        conteo_tip_ped=Detalle_pedido.objects.filter(detallepedido_creado_por=request.user, detallepedido_status=False, detallepedido_tipo_pedido=tipo_pedido).count()
        if conteo_tip_ped > 0:
            pedido=Pedido(pedido_id_depo_id=departamento, pedido_tipoPedido_id=tipo_pedido)
            pedido.save()
            Detalle_pedido.objects.filter(detallepedido_creado_por=request.user, detallepedido_status=False, detallepedido_tipo_pedido=tipo_pedido).update(detallepedido_pedido_id=pedido.pk, detallepedido_status=True)

        return redirect('pedidos:pedido_tienda_listado')

    def post(self, request, *args, **kwargs):
        return HttpResponse('POST request!')


class PedidoList(ListView):
    model = Pedido
    template_name = 'pedidos/pedido/pedido_admin.html'
    ordering = ['pedido_id_pedido']
    def get_context_data(self, **kwargs):
        context = super(PedidoList, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        context['list_sucursal'] = Sucursal.objects.all()
        context['status_list'] = STATUS

        urls_formateada = self.request.GET.copy()
        if 'autoriza' in urls_formateada:
            Pedido.objects.filter(pedido_id_pedido=self.request.GET.get('autoriza')).update(pedido_status=2,pedido_autorizo=self.request.user)
            del urls_formateada['autoriza']

        if 'cancela' in urls_formateada:
            Pedido.objects.filter(pedido_id_pedido=self.request.GET.get('cancela')).update(pedido_status=3,pedido_rechazado=self.request.user)
            del urls_formateada['cancela']
        context['url_format'] = urls_formateada
        print()
        return context

    def get_queryset(self):
        from django.utils.dateparse import parse_datetime
        import pytz
        queryset = super(PedidoList, self).get_queryset()
        status = self.request.GET.get('status')
        inicio = self.request.GET.get('inicio')
        fin = self.request.GET.get('fin')
        sucursal_selec = self.request.GET.get('sucursal_selec')
        busqueda=self.request.GET.get('Buscar')


        if inicio != None and fin != None:
            if inicio != '' and fin != '':
                queryset=Pedido.objects.filter(pedido_fecha_pedido__range=(inicio, fin))

        if sucursal_selec != None:
            if sucursal_selec != '0':
                queryset = queryset.filter(pedido_id_depo__departamento_id_sucursal=sucursal_selec)

        if status != None:
            if status == '0':
                pass
            else:
                queryset = queryset.filter(pedido_status=status)
        else:
            queryset = queryset.filter(pedido_status=1)

        if busqueda != None:
            queryset=Pedido.objects.filter(pedido_id_pedido=busqueda)

        return queryset
    @method_decorator(permission_required('pedidos.view_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PedidoList, self).dispatch(*args, **kwargs)



class DetallePedidoListView(ListView):
    model = Detalle_pedido
    template_name = 'pedidos/pedido/detalle_list.html'

    def get_queryset(self):
        queryset = super(DetallePedidoListView, self).get_queryset()
        queryset = queryset.filter(detallepedido_pedido_id=self.kwargs.get('pk'))
        return queryset

    @method_decorator(permission_required('pedidos.view_detalle_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(DetallePedidoListView, self).dispatch(*args, **kwargs)


class dowload_pedido_detalles(TemplateView):
    def get(self, request , *args, **kwargs):
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        wb = Workbook()
        ws=wb.active
        id_pedido=self.kwargs.get('pk')
        ped = Pedido.objects.get(pedido_id_pedido=id_pedido)
        query = Departamento.objects.get(departamento_id_depo=ped.pedido_id_depo.departamento_id_depo)
        nombre_sucursal = query.departamento_id_sucursal

        ws['A1'] = 'Detalle Pedido '+str(nombre_sucursal)+' N°'+ str(id_pedido)
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Producto'
        ws['B2'] = 'Descripcion'
        ws['C2'] = 'Departamento'
        ws['D2'] = 'Cantidad'
        ws['E2'] = 'Precio'
        ws['F2'] = 'Total'
        cont = 3
        detalle=Detalle_pedido.objects.filter(detallepedido_pedido_id=id_pedido)

        for cto in detalle:
            ws.cell(row=cont, column=1).value = str(cto.detallepedido_producto_id)
            ws.cell(row=cont, column=2).value = str(cto.detallepedido_producto_id.producto_descripcion)
            ws.cell(row=cont, column=3).value = str(cto.detallepedido_pedido_id.pedido_id_depo)
            ws.cell(row=cont, column=4).value = cto.detallepedido_cantidad
            ws.cell(row=cont, column=5).value = cto.detallepedido_precio
            ws.cell(row=cont, column=6).value = (Decimal(cto.detallepedido_cantidad) * Decimal(cto.detallepedido_precio))
            ws.cell(row=cont, column=6).number_format = '#,##0'
            cont += 1

        ws["F"+str(cont)] = "=SUM(F3:F"+str(cont-1)+")"
        ws["F"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        # ACTUALIZAR EL STATUS DE PEDIDO A DESCARGADO
        Pedido.objects.filter(pedido_id_pedido=id_pedido).update(pedido_status=7,pedido_Descargado=request.user)

        nombre_archivo='detalle_pedido_'+str(id_pedido)+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response


class CapturaNoVentaPedido(UpdateView):
    model = Pedido
    form_class = PedidoVentaForm
    template_name = 'pedidos/pedido/pedido_create.html'
    success_url = reverse_lazy('pedidos:pedidos_list')

    def form_valid(self, form):
        form.instance.pedido_status=4
        form.instance.pedido_Venta=self.request.user
        self.object = form.save()
        return super().form_valid(form)
    def get_success_url(self):

        get=self.request.GET.copy()
        url=self.success_url+'?'+get.urlencode()
        return url

    @method_decorator(permission_required('pedidos.change_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CapturaNoVentaPedido, self).dispatch(*args, **kwargs)

class CapturaFacturaPedido(UpdateView):
    model = Pedido
    form_class = PedidoFacturaForm
    template_name = 'pedidos/pedido/pedido_create.html'
    success_url = reverse_lazy('pedidos:pedidos_list')

    def form_valid(self, form):
        form.instance.pedido_status=5
        form.instance.pedido_Facturado=self.request.user
        self.object = form.save()
        return super().form_valid(form)
    def get_success_url(self):

        get=self.request.GET.copy()
        url=self.success_url+'?'+get.urlencode()
        return url

    @method_decorator(permission_required('pedidos.change_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CapturaFacturaPedido, self).dispatch(*args, **kwargs)

class CapturaSalidaPedido(UpdateView):
    model = Pedido
    form_class = PedidoSalidaForm
    template_name = 'pedidos/pedido/pedido_create.html'
    success_url = reverse_lazy('pedidos:pedidos_list')

    def form_valid(self, form):
        form.instance.pedido_status=6
        form.instance.pedido_Finalizado=self.request.user
        self.object = form.save()
        return super().form_valid(form)
    def get_success_url(self):

        get=self.request.GET.copy()
        url=self.success_url+'?'+get.urlencode()
        return url

    @method_decorator(permission_required('pedidos.change_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CapturaSalidaPedido, self).dispatch(*args, **kwargs)

class PedidoUpdate(UpdateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedidos/pedido/pedido_create.html'
    success_url = reverse_lazy('pedidos:pedidos_list')

    def get_context_data(self, **kwargs):
        context = super(PedidoUpdate, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        return context

    def get_success_url(self):
        messages.success(self.request, 'Actualizado Correctamente.')
        # messages.add_message(self.request, messages.ERROR, 'Over 9000!', extra_tags='danger')
        # messages.info(self.request, 'info.')
        # messages.warning(self.request, 'warnin.')
        return reverse_lazy('pedidos:pedido_update', kwargs={'pk':self.kwargs.get('pk')})

    def form_valid(self, form):
        self.object = form.save()
        if form.instance.pedido_status == 2:
            self.object.pedido_autorizo=self.request.user
        elif form.instance.pedido_status == 3:
            self.object.pedido_rechazado=self.request.user
            Detalle_pedido.objects.filter(detallepedido_pedido_id=self.kwargs.get('pk')).update(detallepedido_status=False)
        return super().form_valid(form)

    @method_decorator(permission_required('pedidos.change_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PedidoUpdate, self).dispatch(*args, **kwargs)
                


class PedidoDelete(DeleteView):
    model = Pedido
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:pedidos_list')
    def get_context_data(self, **kwargs):
        context = super(PedidoDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    @method_decorator(permission_required('pedidos.delete_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PedidoDelete, self).dispatch(*args, **kwargs)


class PedidoListSucursal(ListView):
    model = Pedido
    template_name = 'pedidos/pedido/pedido_admin.html'
    ordering = ['pedido_id_pedido']

    def get_context_data(self, **kwargs):
        context = super(PedidoListSucursal, self).get_context_data(**kwargs)
        context['tituloBrea'] = 'Actualizar'
        context['usuario'] = self.request.user
        context['status_list'] = STATUS
        return context

    def get_queryset(self):
        queryset = super(PedidoListSucursal, self).get_queryset()
        try:
            # OBTENER A QUE SUCURSAL PERTENECE EL USUARIO
            pertenece=Pertenece_empresa.objects.get(pertenece_id_usuario=self.request.user)

            status = self.request.GET.get('status')
            inicio = self.request.GET.get('inicio')
            fin = self.request.GET.get('fin')

            if inicio != None and fin != None:
                if inicio != '' and fin != '':
                    queryset=queryset.filter(pedido_id_depo=pertenece.pertenece_empresa, pedido_fecha_pedido__range=(inicio, fin))

            if status != None:
                if status == '0':
                    pass
                else:
                    queryset = queryset.filter(pedido_status=status, pedido_id_depo=pertenece.pertenece_empresa)
            else:
                queryset=queryset.filter(pedido_id_depo=pertenece.pertenece_empresa, pedido_status=1)

        except ObjectDoesNotExist as NoExiste:
            messages.warning(self.request, 'Actual mente usted no pertenece a un departamento, por lo tanto no puede visualizar pedidos realizados.')
            queryset=Pedido.objects.none()

        return queryset

    @method_decorator(permission_required('pedidos.view_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PedidoListSucursal, self).dispatch(*args, **kwargs)

class SelectTipoCompraView(ListView):
    model=Tipo_Pedido
    template_name = "pedidos/select_compra.html"

    def get_queryset(self):
        queryset = super(SelectTipoCompraView, self).get_queryset()
        try:
            object_empresa_user=Pertenece_empresa.objects.get(pertenece_id_usuario=self.request.user)
            queryset = queryset.filter(tp_empresa=object_empresa_user.pertenece_empresa.departamento_id_sucursal.sucursal_empresa_id)
        except ObjectDoesNotExist as error:
            messages.info(self.request, 'No ha sido asignado a una empresa, contacte con el administrador del sitio.')
            queryset=queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        import datetime
        context = super(SelectTipoCompraView, self).get_context_data(**kwargs)
        context['conf'] = Configuracion_pedido.objects.all()
        context['conteo'] = Detalle_pedido.objects.filter(detallepedido_creado_por=self.request.user, detallepedido_status=False).count()
        hoy=datetime.datetime.now()
        for config in context['conf']:
            if config.conf_fecha_inicio <= hoy.date() and config.conf_fecha_fin >= hoy.date():
                context['estado_rango_fechas']=True
        return context

   


class ConfigPedidoListView(ListView):
    model = Configuracion_pedido
    template_name = "pedidos/pedido/config_list.html"

    @method_decorator(permission_required('pedidos.view_configuracion_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ConfigPedidoListView, self).dispatch(*args, **kwargs)

class ConfigPedidoCreate(CreateView):
    model = Configuracion_pedido
    form_class = ConfigForm
    template_name = "pedidos/pedido/conf_create.html"
    success_url = reverse_lazy('pedidos:pedido_config')

    @method_decorator(permission_required('pedidos.add_configuracion_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ConfigPedidoCreate, self).dispatch(*args, **kwargs)

class ConfigPedidoUpdate(UpdateView):
    model = Configuracion_pedido
    form_class = ConfigForm
    template_name = "pedidos/pedido/conf_create.html"
    success_url = reverse_lazy('pedidos:pedido_config')

    @method_decorator(permission_required('pedidos.change_configuracion_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ConfigPedidoUpdate, self).dispatch(*args, **kwargs)
    

# DESCARGA DE DETALLE DE PEDIDOS EN ADMIN FILTRADO

class dowload_report_pedidos(View):
    def get(self, request , *args, **kwargs):
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from django.contrib.humanize.templatetags.humanize import naturalday
        wb = Workbook()
        ws=wb.active

        inicio = self.request.GET.get('inicio')
        fin = self.request.GET.get('fin')
        status = self.request.GET.get('status')
        sucursal_selec = self.request.GET.get('sucursal_selec')

        query_list=Pedido.objects.all()



        if inicio != None and fin != None:
            if status == '0':
                query_list=query_list.filter(pedido_fecha_pedido__range=(inicio, fin))
            else:
                query_list=query_list.filter(pedido_fecha_pedido__range=(inicio, fin), pedido_status=status)
        # else:
        #     if status == '0':
        #         query_list=Pedido.objects.all()
        #     else:
        #         query_list=Pedido.objects.filter(pedido_status=status)

        if sucursal_selec != '0':
            query_list = query_list.filter(pedido_id_depo__departamento_id_sucursal=sucursal_selec)


        ws['A1'] = 'Reporte de pedidos sucursales'
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:M1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '##'
        ws['B2'] = 'CREADO'
        ws['C2'] = 'DEPARTAMENTO'
        ws['D2'] = 'SUCURSAL'
        ws['E2'] = 'SUBTOTAL'
        ws['F2'] = 'AUTORIZO'
        ws['G2'] = 'RECHAZÓ'
        ws['H2'] = 'Factura'
        ws['I2'] = 'Venta'
        ws['J2'] = 'STATUS'
        ws['K2'] = 'TIPO'
        cont = 3
        for cto in query_list:

            ws.cell(row=cont, column=1).value = str(cto.pedido_id_pedido)
            ws.cell(row=cont, column=2).value = str(cto.pedido_fecha_pedido)
            ws.cell(row=cont, column=3).value = str(cto.pedido_id_depo)
            ws.cell(row=cont, column=4).value = cto.pedido_id_depo.departamento_nombre
            # ws.cell(row=cont, column=5).value = cto.get_total_tipo_pedido()
            ws.cell(row=cont, column=5).value = cto.get_total()
            ws.cell(row=cont, column=6).value = 'NO' if cto.pedido_autorizo == None else cto.pedido_autorizo.username
            ws.cell(row=cont, column=7).value = 'NO' if cto.pedido_rechazado == None else cto.pedido_rechazado.username
            ws.cell(row=cont, column=8).value = '' if cto.pedido_n_factura == None else cto.pedido_n_factura
            ws.cell(row=cont, column=9).value = '' if cto.pedido_n_cresscedo == None else cto.pedido_n_cresscedo
            ws.cell(row=cont, column=10).value = cto.get_pedido_status_display()
            ws.cell(row=cont, column=11).value = str(cto.pedido_tipoPedido)
            ws.cell(row=cont, column=5).number_format = '#,##0.00'
            cont += 1

        ws.cell(row=cont, column=4).value = 'TOTAL'
        ws["E"+str(cont)] = "=SUM(E3:E"+str(cont-1)+")"
        ws["E"+str(cont)].number_format = '#,##0.00'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+2



        nombre_archivo='report_sucursales'+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response


class DowloadDetallesporPedido(View):
    def get(self, request , *args, **kwargs):
        import random
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from openpyxl.styles import PatternFill
        wb = Workbook()
        ws=wb.active

        inicio=request.GET.get('inicio')
        fin=request.GET.get('fin')
        sucursal_selec=request.GET.get('sucursal_selec')
        status=request.GET.get('status')

        query = Detalle_pedido.objects.filter(detallepedido_pedido_id__pedido_fecha_pedido__range=(inicio, fin))

        if sucursal_selec != '0':
            query=query.filter(detallepedido_pedido_id__pedido_id_depo__departamento_id_sucursal=sucursal_selec)
        if status != '0':
            query=query.filter(detallepedido_pedido_id__pedido_status=status)

        query=query.order_by('detallepedido_pedido_id')
        # query=query.order_by('detallepedido_pedido_id__pedido_tipo')

        ws['A1'] = 'Detalle pedidos filtrado'
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '#Pedido'
        ws['B2'] = 'Departamento'
        ws['C2'] = 'Sucursal'
        ws['D2'] = 'Tipo Ped'
        ws['E2'] = 'Producto'
        ws['F2'] = 'Descripcion'
        ws['G2'] = 'Cantidad'
        ws['H2'] = 'Precio'
        ws['I2'] = 'Subtotal'
        cont = 3
        colors_rbga=['EFA8D3', 'BEABF6', 'F3B8F8', '6FE3F5', 'F4B2B2', 'FFB7D5', 'A7ABFA', '94F4B5', 'F7D896', 'D4D3D3', 'FFBCBC', 'F8CEA4', 'CEF683', 'C8FBC0', '87C8E2', 'A7C4FA', 'F49495', 'D4B1EB']
        tem_cont_ped=0


        for cto in query:
            # CONDICIONAL PARA CAMBIO DE COLOR
            if cto.detallepedido_pedido_id.pedido_id_pedido != tem_cont_ped:
                color_temporal_gen=random.choice(colors_rbga)
            tem_cont_ped=cto.detallepedido_pedido_id.pedido_id_pedido
            # COLOREAR CELDAS
            ws.cell(row=cont, column=1).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=2).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=3).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=4).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=5).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=6).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=7).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=8).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=cont, column=9).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            # INCRUSTAR CONTENIDO DE CELDA
            ws.cell(row=cont, column=1).value = cto.detallepedido_pedido_id.pedido_id_pedido
            ws.cell(row=cont, column=2).value = str(cto.detallepedido_pedido_id.pedido_id_depo.departamento_nombre)
            ws.cell(row=cont, column=3).value = str(cto.detallepedido_pedido_id.pedido_id_depo.departamento_id_sucursal)
            ws.cell(row=cont, column=4).value = str(cto.detallepedido_pedido_id.pedido_tipoPedido)
            ws.cell(row=cont, column=5).value = str(cto.detallepedido_producto_id)
            ws.cell(row=cont, column=6).value = str(cto.detallepedido_producto_id.producto_nombre)
            ws.cell(row=cont, column=7).value = cto.detallepedido_cantidad
            ws.cell(row=cont, column=8).value = cto.detallepedido_precio
            ws.cell(row=cont, column=9).value = '=PRODUCT(G'+str(cont)+':H'+str(cont)+')'
            ws.cell(row=cont, column=9).number_format = '#,##0.00'
            cont += 1

        ws["I"+str(cont)] = "=SUM(I3:I"+str(cont-1)+")"
        ws["I"+str(cont)].number_format = '#,##0.00'


        # CODIGO PARA AJUSTAR LAS CELDAS EN EL EXCEL
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1



        nombre_archivo='detalle_pedido_filtrado.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

class TipoPedidoList(ListView):
    template_name = 'pedidos/tipo_pedido/tipopedido_list.html'
    model=Tipo_Pedido

    @method_decorator(permission_required('pedidos.view_tipo_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
         return super(TipoPedidoList, self).dispatch(*args, **kwargs)

class TipoPedidoCrear(CreateView):
    model=Tipo_Pedido
    template_name = 'pedidos/tipo_pedido/tipo_pedido_crear.html'
    form_class=Tipo_PedidoForm
    success_url = reverse_lazy('pedidos:config_tipo_pedido_list')

    @method_decorator(permission_required('pedidos.add_tipo_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(TipoPedidoCrear, self).dispatch(*args, **kwargs)

class TipoPedidoEdit(UpdateView):
    model=Tipo_Pedido
    template_name = 'pedidos/tipo_pedido/tipo_pedido_crear.html'
    form_class=Tipo_PedidoForm
    success_url = reverse_lazy('pedidos:config_tipo_pedido_list')

    @method_decorator(permission_required('pedidos.change_tipo_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(TipoPedidoEdit, self).dispatch(*args, **kwargs)

class TipoPedidoDelete(DeleteView):
    model = Tipo_Pedido
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:config_tipo_pedido_list')
    def get_context_data(self, **kwargs):
        context = super(TipoPedidoDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    @method_decorator(permission_required('pedidos.delete_tipo_pedido',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(TipoPedidoDelete, self).dispatch(*args, **kwargs)

class AsigGastoList(ListView):
    model=Asignar_gasto_sucursal
    template_name = 'pedidos/tipo_pedido/asignar_gasto_list.html'
    def get_context_data(self, **kwargs):
        import datetime
        context = super(AsigGastoList, self).get_context_data(**kwargs)
        context['sucursal_list'] = Sucursal.objects.all()
        context['tipo_pedido_list'] = Tipo_Pedido.objects.all()
        return context

    def get_queryset(self):
        queryset = super(AsigGastoList, self).get_queryset()
        tip_pedido=self.request.GET.get('tip_pedido')
        sucursal=self.request.GET.get('sucursal')
        if tip_pedido != None:
            queryset=queryset.filter(ags_tipo_ped=tip_pedido)
        if sucursal != None:
            queryset=queryset.filter(ags_sucursal__departamento_id_sucursal=sucursal)
        return queryset

    @method_decorator(permission_required('pedidos.view_asignar_gasto_sucursal',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
         return super(AsigGastoList, self).dispatch(*args, **kwargs)


class AsigGastoCrear(CreateView):
    model=Asignar_gasto_sucursal
    template_name = 'pedidos/tipo_pedido/tipo_pedido_crear.html'
    form_class=AsigGastoForm
    success_url = reverse_lazy('pedidos:asig_gasto_list')

    @method_decorator(permission_required('pedidos.add_asignar_gasto_sucursal',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(AsigGastoCrear, self).dispatch(*args, **kwargs)

class AsigGastoUpdate(UpdateView):
    model=Asignar_gasto_sucursal
    template_name = 'pedidos/tipo_pedido/tipo_pedido_crear.html'
    form_class=AsigGastoForm
    success_url = reverse_lazy('pedidos:asig_gasto_list')

    @method_decorator(permission_required('pedidos.change_asignar_gasto_sucursal',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(AsigGastoUpdate, self).dispatch(*args, **kwargs)

class AsigGastoDelete(DeleteView):
    model = Asignar_gasto_sucursal
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('pedidos:asig_gasto_list')

    def get_context_data(self, **kwargs):
        context = super(AsigGastoDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context 

    @method_decorator(permission_required('pedidos.delete_asignar_gasto_sucursal',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(AsigGastoDelete, self).dispatch(*args, **kwargs)


# CLASES PARA GENERAR CATALOGO DE PEDIDOS

class CatalogoProductosList(ListView):
    model = Catalogo_Productos
    template_name = 'pedidos/catalogo/listado_catalogo.html'

    @method_decorator(permission_required('pedidos.view_catalogo_productos',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
         return super(CatalogoProductosList, self).dispatch(*args, **kwargs)

class CatalogoCreate(CreateView):
    model = Catalogo_Productos
    form_class = Catalogo_ProductosForm
    template_name ='pedidos/catalogo/create_catalogo.html'
    success_url = reverse_lazy('pedidos:listar_catalogo')

    @method_decorator(permission_required('pedidos.add_catalogo_productos',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CatalogoCreate, self).dispatch(*args, **kwargs)

class CatalogoActualizar(UpdateView):
    model = Catalogo_Productos
    form_class = Catalogo_ProductosForm
    template_name ='pedidos/catalogo/create_catalogo.html'
    success_url = reverse_lazy('pedidos:listar_catalogo')

    @method_decorator(permission_required('pedidos.change_catalogo_productos',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CatalogoActualizar, self).dispatch(*args, **kwargs)

class CatalogoDelete(DeleteView):
    model = Catalogo_Productos
    template_name = 'pedidos/catalogo/delete_catalogo.html'
    success_url = reverse_lazy('pedidos:listar_catalogo')

    @method_decorator(permission_required('pedidos.delete_catalogo_productos',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CatalogoDelete, self).dispatch(*args, **kwargs)


class PDFCatalogoProd(View):
    object_catalogo=None
    def myFirstPage(self, canvas, doc):
        from django.utils.formats import localize
        from datetime import datetime
        print('soy lo encabezado')
        #CABECERA DE PAGINA
        Title = "CATALOGO DE PRODUCTOS"
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)

        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 50, Title)       
        #Logo de empresa
        archivo_imagen = self.object_catalogo.tp_empresa.empresa_logo.path
        #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
        canvas.drawImage(archivo_imagen, 20, 690, 120, 90, preserveAspectRatio=True)
       

        canvas.saveState()
        canvas.setFont('Times-Roman', 10)
        
        page_count = doc.page 
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 770, "Página {}".format(page_count))
        canvas.restoreState()

        canvas.setFont('Times-Roman', 10)
        
        canvas.drawString(420, 30, '{}'.format(localize(datetime.now())))


    def myLaterPages(self, canvas, doc):
        print('pie de pagina')
        canvas.saveState()
        canvas.setFont('Times-Roman', 10)
        
        page_count = doc.page 
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 770, "Página {}".format(page_count))
        canvas.restoreState()


    
    def dispatch(self, *args, **kwargs):
        id_obj=self.kwargs.get('pk')
        self.object_catalogo=Catalogo_Productos.objects.get(id=id_obj)
        return super().dispatch(*args, **kwargs)



    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/pdf')
        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=18, title='CATALOGO PRODUCTOS')
        items = []
        data_tabla=[]

        stylo_p_center = ParagraphStyle('parrafo_center', alignment=TA_CENTER, fontSize=11, fontName="Times-Roman")
        stylo_p = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=11, fontName="Times-Roman")
        stylo_titulo = ParagraphStyle('titulo', alignment=TA_CENTER, fontSize=11, fontName="Times-Bold")
        stylo_portada_title = ParagraphStyle('titulo', alignment=TA_CENTER, fontSize=20, fontName="Times-Bold")

        items.append(Image(self.object_catalogo.tp_imagen.path, 8*cm, 8*cm))
        items.append(Paragraph(self.object_catalogo.tp_catalogo, stylo_portada_title))
        items.append(Spacer(0,30))
        items.append(Paragraph(self.object_catalogo.tp_no_licitacion, stylo_portada_title))
        items.append(Spacer(0,30))
        items.append(Paragraph(self.object_catalogo.tp_descripcion, stylo_portada_title))
        items.append(PageBreak())

        
        dta=[]
        if self.object_catalogo.tp_orientacion_t == 1:
            titulos_tabla = [(Paragraph('##', stylo_titulo),Paragraph('Imagen', stylo_titulo), Paragraph('Descripción', stylo_titulo))]
            for item, iteracion in zip(self.object_catalogo.tp_productos.all(), range(self.object_catalogo.tp_productos.all().count())):
                dta.append((iteracion+1,Image(item.producto_imagen.path, 5*cm, 5*cm), item.producto_descripcion))
            tabla=Table(titulos_tabla+dta, colWidths=[1*cm, 6 * cm, 13 * cm])
        elif self.object_catalogo.tp_orientacion_t == 2:
            titulos_tabla = [(Paragraph('##', stylo_titulo),Paragraph('Descripción', stylo_titulo),Paragraph('Imagen', stylo_titulo))]
            for item, iteracion in zip(self.object_catalogo.tp_productos.all(), range(self.object_catalogo.tp_productos.all().count())):
                dta.append((iteracion+1, item.producto_descripcion, Image(item.producto_imagen.path, 5*cm, 5*cm)))
            tabla=Table(titulos_tabla+dta, colWidths=[1*cm, 13 * cm, 6 * cm])

        tabla.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.dodgerblue),
                # ('LINEBELOW', (0, 0), (-1, 0), 0, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.transparent)
            ]
        ))
        items.append(tabla)
        items.append(Spacer(0,20))


        doc.build(items, onFirstPage=self.myFirstPage, onLaterPages = self.myLaterPages)
        response.write(buff.getvalue())
        buff.close()
        return response




class InventarioBusqueda(TemplateView):
    template_name='pedidos/inventario/busqueda.html'
    def get_context_data(self, **kwargs):
        context = super(InventarioBusqueda, self).get_context_data(**kwargs)
        buscar=self.request.GET.get('Buscar')
        if buscar != '':
            if buscar != None:
                context['produsctos'] = Producto.objects.filter(Q(producto_descripcion__icontains=buscar) | Q(producto_codigo__icontains=buscar) | Q(prducto_codigo_barras__icontains=buscar), producto_visible=False)
        return context


class InvCapturaResguardo(CreateView):
    model = Inventario
    template_name = 'pedidos/inventario/formulario.html'
    form_class = InventarioResguardoForm
    success_url = reverse_lazy('pedidos:inventarion_look_up')

    def form_valid(self, form):
        try:
            form.instance.inv_user_catura=self.request.user
            form.instance.inv_tipo_sitio=1
            form.instance.inv_producto=Producto.objects.get(producto_codigo=form.cleaned_data['inv_producto'])
            self.object = form.save()
            messages.success(self.request, 'Agregado')
            return super().form_valid(form)
        except IntegrityError as error:
            messages.warning(self.request, 'Producto y ubicacion duplicado intente con otra ubicacion'+str(error))
            url=reverse_lazy('pedidos:inventarion_look_up')
            url=url+'?Buscar='+form.cleaned_data['inv_producto']
            return redirect(url)
        

    @method_decorator(permission_required('pedidos.add_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InvCapturaResguardo, self).dispatch(*args, **kwargs)

class InvCapturaPikin(CreateView):
    model = Inventario
    template_name = 'pedidos/inventario/formulario.html'
    form_class = InventarioPikinForm
    success_url = reverse_lazy('pedidos:inventarion_look_up')

    def form_valid(self, form):
        try:
            form.instance.inv_user_catura=self.request.user
            form.instance.inv_tipo_sitio=2
            form.instance.inv_producto=Producto.objects.get(producto_codigo=form.cleaned_data['inv_producto'])
            self.object = form.save()
            messages.success(self.request, 'Agregado')
            return super().form_valid(form)
        except IntegrityError as error:
            messages.warning(self.request, 'Producto y ubicacion duplicado intente con otra ubicacion'+str(error))
            url=reverse_lazy('pedidos:inventarion_look_up')
            url=url+'?Buscar='+form.cleaned_data['inv_producto']
            return redirect(url)
        

    @method_decorator(permission_required('pedidos.add_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InvCapturaPikin, self).dispatch(*args, **kwargs)


class InvCapturaOtros(CreateView):
    model = Inventario
    template_name = 'pedidos/inventario/formulario.html'
    form_class = InventarioOtrosForm
    success_url = reverse_lazy('pedidos:inventarion_look_up')

    def form_valid(self, form):
        try:
            form.instance.inv_user_catura=self.request.user
            form.instance.inv_tipo_sitio=3
            form.instance.inv_producto=Producto.objects.get(producto_codigo=form.cleaned_data['inv_producto'])
            self.object = form.save()
            messages.success(self.request, 'Agregado')
            return super().form_valid(form)
        except IntegrityError as error:
            messages.warning(self.request, 'Producto y ubicacion duplicado intente con otra ubicacion'+str(error))
            url=reverse_lazy('pedidos:inventarion_look_up')
            url=url+'?Buscar='+form.cleaned_data['inv_producto']
            return redirect(url)
        

    @method_decorator(permission_required('pedidos.add_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InvCapturaOtros, self).dispatch(*args, **kwargs)


class InvCapturaMerma(CreateView):
    model = Inventario
    template_name = 'pedidos/inventario/formulario.html'
    form_class = InventarioMermaForm
    success_url = reverse_lazy('pedidos:inventarion_look_up')

    def form_valid(self, form):
        try:
            form.instance.inv_user_catura=self.request.user
            form.instance.inv_validacion=False
            form.instance.inv_tipo_sitio=4
            form.instance.inv_producto=Producto.objects.get(producto_codigo=form.cleaned_data['inv_producto'])
            self.object = form.save()
            messages.success(self.request, 'Agregado')
            return super().form_valid(form)
        except IntegrityError as error:
            messages.warning(self.request, 'Producto y ubicacion duplicado intente con otra ubicacion'+str(error))
            url=reverse_lazy('pedidos:inventarion_look_up')
            url=url+'?Buscar='+form.cleaned_data['inv_producto']
            return redirect(url)
        

    @method_decorator(permission_required('pedidos.add_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InvCapturaMerma, self).dispatch(*args, **kwargs)


class InventarioAvanceConteo(ListView):
    model=Inventario
    paginate_by=300
    template_name='pedidos/inventario/avance.html'

    @method_decorator(permission_required('pedidos.change_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InventarioAvanceConteo, self).dispatch(*args, **kwargs)

    def get_queryset(self):
        queryset = super(InventarioAvanceConteo, self).get_queryset()
        buscar=self.request.GET.get('Buscar')
        cod_prod=self.request.GET.get('cod_prod')

        if cod_prod != None:
            queryset=queryset.filter(inv_producto=cod_prod)

        if buscar != '':
            if buscar != None:
                queryset = queryset.filter(Q(inv_producto__producto_descripcion__icontains=buscar) | Q(inv_producto__producto_codigo__icontains=buscar), inv_producto__producto_visible=False)[:100]
        return queryset
    

class InventarioAvanceConteoGlobal(ListView):
    model=Inventario
    paginate_by = 500
    template_name='pedidos/inventario/Global.html'
    def get_queryset(self):
        queryset = super(InventarioAvanceConteoGlobal, self).get_queryset()
        tipo_conteo=self.request.GET.get('conteo')
        queryset=Inventario.objects.values('inv_producto', 'inv_producto__producto_descripcion', 'inv_producto__prducto_existencia').filter(inv_producto__producto_visible=False,inv_validacion=True, inv_conteo=tipo_conteo).annotate(
            total_resguardo=Sum('inv_cant_resguardo'), 
            total_piking=Sum('inv_cant_piking'), 
            total_otros=Sum('inv_cant_otros'), 
            total_merma=Sum('inv_cant_merma'), 
            suma_total=Sum(F('inv_cant_resguardo'))+Sum(F('inv_cant_piking'))+Sum(F('inv_cant_otros')),
            ).order_by('inv_producto')        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super(InventarioAvanceConteoGlobal, self).get_context_data(**kwargs)
        tipo_conteo=self.request.GET.get('conteo')
        context['total_conteo_prod'] = Inventario.objects.values('inv_producto').filter(inv_producto__producto_visible=False,inv_validacion=True, inv_conteo=tipo_conteo).annotate(
            total_resguardo=Sum('inv_cant_resguardo'),
            ).order_by('inv_producto').count()
        context['total_productos']=Producto.objects.filter(producto_visible=False).count()
        avance_porcentaje=(context['total_conteo_prod']*100)/context['total_productos']
        context['avance_porcentaje']=round(avance_porcentaje, 2)
        context['conteo']=CONTEO
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context


class ValidarConteoMerma(TemplateView):
    template_name = 'pedidos/inventario/validar.html'

    def get_context_data(self, **kwargs):
        context = super(ValidarConteoMerma, self).get_context_data(**kwargs)
        context['inventario'] = Inventario.objects.filter(inv_sup_autorizo_merma=self.request.user, inv_validacion=False)
        id_inv=self.request.GET.get('autorizo')
        id_rechazar=self.request.GET.get('rechazar')
        if id_inv != None:
            Inventario.objects.filter(inv=id_inv).update(inv_validacion=True)

        if id_rechazar != None:
            Inventario.objects.filter(inv=id_rechazar).delete()
        
        return context


class InventarioDelete(DeleteView):
    model=Inventario
    template_name="pedidos/delete_forever.html"
    success_message = "Eliminado correctamente."
    success_url=reverse_lazy('pedidos:inventarion_avance_conteo')
    @method_decorator(permission_required('pedidos.delete_inventario',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(InventarioDelete, self).dispatch(*args, **kwargs)
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, self.success_message)
        return super(InventarioDelete, self).delete(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super(InventarioDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context
    





class DowloadInventarioGlobal(View):
    def get(self, request , *args, **kwargs):
        import random
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from openpyxl.styles import PatternFill
        wb = Workbook()
        ws=wb.active

        ws['A1'] = 'INVENTARIO CONTEO GLOBAL N°'+self.request.GET.get('conteo')
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Codigo Producto'
        ws['B2'] = 'Descripción'
        ws['C2'] = 'Resguardo'
        ws['D2'] = 'Piking'
        ws['E2'] = 'Otros'
        ws['F2'] = 'Merma'
        ws['G2'] = 'Total'
        ws['H2'] = 'Existencia'
        cont = 3
        tipo_conteo=self.request.GET.get('conteo')
        query=Inventario.objects.values('inv_producto', 'inv_producto__producto_descripcion', 'inv_producto__prducto_existencia').filter(inv_producto__producto_visible=False,inv_validacion=True, inv_conteo=tipo_conteo).annotate(
            total_resguardo=Sum('inv_cant_resguardo'), 
            total_piking=Sum('inv_cant_piking'), 
            total_otros=Sum('inv_cant_otros'), 
            total_merma=Sum('inv_cant_merma'), 
            suma_total=Sum(F('inv_cant_resguardo'))+Sum(F('inv_cant_piking'))+Sum(F('inv_cant_otros')),
            ).order_by('inv_producto')
        


        for cto in query:
            # INCRUSTAR CONTENIDO DE CELDA
            ws.cell(row=cont, column=1).value = cto['inv_producto']
            ws.cell(row=cont, column=2).value = cto['inv_producto__producto_descripcion']
            ws.cell(row=cont, column=3).value = cto['total_resguardo']
            ws.cell(row=cont, column=4).value = cto['total_piking']
            ws.cell(row=cont, column=5).value = cto['total_otros']
            ws.cell(row=cont, column=6).value = cto['total_merma']
            ws.cell(row=cont, column=7).value = cto['suma_total']
            ws.cell(row=cont, column=8).value = cto['inv_producto__prducto_existencia']

            ws.cell(row=cont, column=3).number_format = '#,##0'
            ws.cell(row=cont, column=4).number_format = '#,##0'
            ws.cell(row=cont, column=5).number_format = '#,##0'
            ws.cell(row=cont, column=6).number_format = '#,##0'
            ws.cell(row=cont, column=7).number_format = '#,##0'
            ws.cell(row=cont, column=8).number_format = '#,##0'
            cont += 1


        # CODIGO PARA AJUSTAR LAS CELDAS EN EL EXCEL
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1



        nombre_archivo='conteo_global_inventario_no_'+tipo_conteo+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

class DowloadInventarioGlobalComparativo(View):
    def get(self, request , *args, **kwargs):
        import random
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from openpyxl.styles import PatternFill
        wb = Workbook()
        ws=wb.active

        ws['A1'] = 'INVENTARIO CONTEO GLOBAL COMPARADO'
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Codigo Producto'
        ws['B2'] = 'Descripción'
        ws['C2'] = 'Resguardo'
        ws['D2'] = 'Piking'
        ws['E2'] = 'Otros'
        ws['F2'] = 'Merma'
        ws['G2'] = 'Total'
        ws['H2'] = 'Existencia'
        ws['I2'] = 'N° Conteo'
        cont = 3
        
        query=Inventario.objects.values('inv_producto', 'inv_producto__producto_descripcion', 'inv_producto__prducto_existencia', 'inv_conteo').filter(inv_producto__producto_visible=False,inv_validacion=True).annotate(
            total_resguardo=Sum('inv_cant_resguardo'), 
            total_piking=Sum('inv_cant_piking'), 
            total_otros=Sum('inv_cant_otros'), 
            total_merma=Sum('inv_cant_merma'), 
            suma_total=Sum(F('inv_cant_resguardo'))+Sum(F('inv_cant_piking'))+Sum(F('inv_cant_otros')),
            ).order_by('inv_producto', 'inv_conteo')
        


        for cto in query:
            # INCRUSTAR CONTENIDO DE CELDA
            ws.cell(row=cont, column=1).value = cto['inv_producto']
            ws.cell(row=cont, column=2).value = cto['inv_producto__producto_descripcion']
            ws.cell(row=cont, column=3).value = cto['total_resguardo']
            ws.cell(row=cont, column=4).value = cto['total_piking']
            ws.cell(row=cont, column=5).value = cto['total_otros']
            ws.cell(row=cont, column=6).value = cto['total_merma']
            ws.cell(row=cont, column=7).value = cto['suma_total']
            ws.cell(row=cont, column=8).value = cto['inv_producto__prducto_existencia']
            ws.cell(row=cont, column=9).value = cto['inv_conteo']

            ws.cell(row=cont, column=3).number_format = '#,##0'
            ws.cell(row=cont, column=4).number_format = '#,##0'
            ws.cell(row=cont, column=5).number_format = '#,##0'
            ws.cell(row=cont, column=6).number_format = '#,##0'
            ws.cell(row=cont, column=7).number_format = '#,##0'
            ws.cell(row=cont, column=8).number_format = '#,##0'
            ws.cell(row=cont, column=9).number_format = '#,##0'
            cont += 1


        # CODIGO PARA AJUSTAR LAS CELDAS EN EL EXCEL
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1



        nombre_archivo='conteo_global_inventario_comparativo.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

class ResetInventarioDelete(TemplateView):
    template_name="pedidos/delete_forever.html"
    def post(self, request, *args, **kwargs):
        Inventario.objects.all().delete()
        return redirect('pedidos:inventarion_avance_conteo_global')

    def get_context_data(self, **kwargs):
        context = super(ResetInventarioDelete, self).get_context_data(**kwargs)
        query=Inventario.objects.all()
        deletable_objects, model_count, protected = get_deleted_objects(query)
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context 