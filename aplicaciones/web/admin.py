from django.contrib import admin
from django import forms
from aplicaciones.web.forms import BlogForms
from aplicaciones.web.models import *
from django.contrib import messages
class ConfigCorreos(admin.ModelAdmin):
    list_display = ('corr_nombre', 
                    'corr_email',
                    'corr_telefono',
                    'corr_asunto',
                    'corr_mensaje',
                    'corr_depo',
                    'corr_crado',)
    list_filter = ['corr_depo', 'corr_crado']
class ConfigRegistro(admin.ModelAdmin):
    list_display = ('re_nombre', 
                    're_apellidos',
                    're_correo',
                    're_telefono',
                    're_evento',
                    're_creado',
                    )
    list_filter = ['re_evento', 're_creado']


class ConfigPostulate(admin.ModelAdmin):
    list_display = ('pos_vacante', 
                    'pos_nombre',
                    'pos_correo',
                    'pos_telefono',
                    'pos_cv',
                    'pos_creacion',
                    )
    list_filter = ['pos_vacante', 'pos_creacion']



class DetalleCompraWebInline(admin.TabularInline):
    model = Detalle_Compra_Web
    readonly_fields = ('dcw_precio', )
    raw_id_fields=('dcw_producto_id', )
    # Mostramos dos inlines acíos por defecto
    extra = 0
    
class ConfigCompraWeb(admin.ModelAdmin):
    
    actions = ['dowload_xls', 'mark_entrega_pedido', 'delete_selected']
    list_display = ('cw_id', 
                    'cw_fecha',
                    'cliente_str',
                    'cw_status',
                    'domicilio',
                    'cw_numero_venta',
                    'cw_numero_factura',
                    'cw_tipo_pago',
                    )
    list_filter = ['cw_fecha', 'cw_status']
    search_fields = ['cw_id', 'cw_cliente__rfc', 'cw_cliente__username']
    readonly_fields = ('domicilio', 'cw_tipo_pago', 'cw_status', 'cliente_str')
    exclude = ('cw_domicilio','cw_cliente')
    inlines = [DetalleCompraWebInline]

    def save_model(self, request, obj, form, change):        
        if change:
            if obj.cw_status == 2:
                if form.instance.cw_numero_venta != None:
                    obj.cw_status=3

            if obj.cw_status == 3:
                if form.instance.cw_numero_factura != None:
                    obj.cw_status=4            
        super().save_model(request, obj, form, change)

    def mark_entrega_pedido(self, request, queryset):
        conteo=queryset.filter(cw_status=4).update(cw_status=5)
        self.message_user(request, '{} Compras han sido marcadas como entregadas.'.format(conteo))
    mark_entrega_pedido.short_description = "Marcar como entregado"
    mark_entrega_pedido.allowed_permissions = ('change',)

    def dowload_xls(self, request, queryset):
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from django.contrib.humanize.templatetags.humanize import naturalday
        wb = Workbook()
        ws = wb.active

        

        ws['A1'] = 'PEDIDOS WEB'
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws.sheet_properties.tabColor = "1072BA"

        
        cont = 2
        for item in queryset:
            ws['A{}'.format(cont)] = 'Pedido N° {}'.format(item.cw_id)
            st = ws['A{}'.format(cont)]
            st.font = Font(size=12, b=True, color="00e0e0")
            st.alignment = Alignment(horizontal='center')
            ws.merge_cells('A{}:E{}'.format(cont, cont))
            cont += 1
            ws['A{}'.format(cont)] = '##'
            ws['B{}'.format(cont)] = 'PRODUCTO'
            ws['C{}'.format(cont)] = 'CANTIDAD' 
            ws['D{}'.format(cont)] = 'PRECIO'
            ws['E{}'.format(cont)] = 'SUBTOTAL'
            cont += 1
            for detalle in Detalle_Compra_Web.objects.filter(dcw_pedido_id=item.cw_id):
                ws.cell(row=cont, column=1).value = str(detalle.dcw_pedido_id)
                ws.cell(row=cont, column=2).value = str(detalle.dcw_producto_id)
                ws.cell(row=cont, column=3).value = detalle.dcw_cantidad
                ws.cell(row=cont, column=4).value = detalle.dcw_precio
                ws.cell(row=cont, column=5).value = (detalle.dcw_cantidad*detalle.dcw_precio)
                ws.cell(row=cont, column=5).number_format = '#,##0.00'
                cont += 1
            
            
        
        queryset.filter(cw_status=1).update(cw_status=2)

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+2

        nombre_archivo = 'pedidos_web'+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    dowload_xls.short_description = "Descargar xls"
    dowload_xls.allowed_permissions = ('change',)






class BlogConfig(admin.ModelAdmin):
    form = BlogForms
    filter_horizontal = ('blog_tags',)
    actions = ['activar_portada', 'desactivar_portadas', 'activar_oferta_semanal', 'activar_oferta_especial', 'activar_banner']
    list_display = (
        'blog_titulo',
        'blog_creado',
        'blog_pertenece',
        'blog_categoria',
        'blog_tipo',        
        )
    # list_editable = ('blog_tipo',)
    def save_model(self, request, obj, form, change):
        obj.blog_pertenece = request.user
        super().save_model(request, obj, form, change)
    
    def desactivar_portadas(self, request, queryset):
        port=queryset.update(blog_tipo=1)
        messages.success(request, 'Portada desactivadas {}.'.format(port))
    desactivar_portadas.short_description = "Desactivar portada y/o ofertas" 

    def activar_portada(self, request, queryset):
        port=queryset.update(blog_tipo=2)
        messages.success(request, 'Portada activadas {}.'.format(port))
    activar_portada.short_description = "Activar como portada"

    def activar_oferta_semanal(self, request, queryset):
        port=queryset.update(blog_tipo=3)
        messages.success(request, 'Ofertas semanal {}.'.format(port))
    activar_oferta_semanal.short_description = "Activar oferta semanal"

    def activar_oferta_especial(self, request, queryset):
        port=queryset.update(blog_tipo=4)
        messages.success(request, 'Oferta especial {}.'.format(port))
    activar_oferta_especial.short_description = "Activar oferta especial"

    def activar_banner(self, request, queryset):
        port=queryset.update(blog_tipo=5)
        messages.success(request, 'Banner en pagina de compra {}.'.format(port))
    activar_banner.short_description = "Banner en pagina de compra"

    



admin.site.register(Departamento)
admin.site.register(CorreoCco, ConfigCorreos)
admin.site.register(Evento)
admin.site.register(RegistroExpo, ConfigRegistro)
admin.site.register(Marca)
admin.site.register(Catalagos)
admin.site.register(Promocion)

admin.site.register(Vacante)
admin.site.register(Postulacion, ConfigPostulate)
admin.site.register(CompraWeb, ConfigCompraWeb)
admin.site.register(Tag)
admin.site.register(Blog, BlogConfig)
admin.site.register(DescuentoTotal)
