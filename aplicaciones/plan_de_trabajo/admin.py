from django.contrib import admin
from aplicaciones.plan_de_trabajo.models import Plan_trabajo, Registro_actividad, VisitaVendedor
from aplicaciones.empresa.models import Cliente
from aplicaciones.plan_de_trabajo.forms import Registro_actividadForm
from django.db.models import Sum, F, Count, Max, Min


from rangefilter.filter import DateRangeFilter, DateTimeRangeFilter

import datetime
from django.utils import timezone
from django.shortcuts import redirect
from aplicaciones.inicio.models import User
Usuario = User
# Register your models here.

class UserStafFilter(admin.SimpleListFilter):
    title = ('Vendedor')
    parameter_name = 'pt_vendedor'
    def lookups(self, request, model_admin):
       """
           List of values to allow admin to select
       """
       usr = Usuario.objects.filter(is_staff=True)
       lista=[]
       for item in usr:
           lista.append((item.id,(item.username)))
       return lista
    def queryset(self, request, queryset):
        if self.value() == None:
            queryset=queryset.all()
        else:
            queryset=queryset.filter(pt_vendedor=self.value())
        return queryset

class UserStafActividadFilter(admin.SimpleListFilter):
    title = ('Vendedor')
    parameter_name = 'ra_user'
    def lookups(self, request, model_admin):
       """
           List of values to allow admin to select
       """
       usr = Usuario.objects.filter(is_staff=True)
       lista=[]
       for item in usr:
           lista.append((item.id,(item.username)))
       return lista
    def queryset(self, request, queryset):
        if self.value() == None:
            queryset=queryset.all()
        else:
            queryset=queryset.filter(ra_user=self.value())
        return queryset

class Plan_trabajoAdmin(admin.ModelAdmin):
    list_display=[
        'pt', 
        'pt_cliente', 
        'pt_vendedor', 
        'pt_lunes', 
        'pt_martes', 
        'pt_miercoles', 
        'pt_jueves', 
        'pt_viernes', 
        'pt_sabado', 
        'pt_domingo', 
        ]
    list_filter = [UserStafFilter]
    search_fields = ['pt_vendedor__username', 'pt']
    raw_id_fields = ('pt_cliente',)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if form.base_fields:
            form.base_fields['pt_vendedor'].queryset = Usuario.objects.filter(is_staff=True)
        return form
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.has_perm('plan_de_trabajo.reg_activ_supervisor'):
            qs=qs.all()
        else:
            qs=qs.filter(pt_vendedor=request.user)
        return qs




class AdminRegistroActividad(admin.ModelAdmin):
    exclude = ('ra_user',)
    actions = ['DownloadReportSelected']
    list_filter = [('ra_fecha_creacion', DateRangeFilter), UserStafActividadFilter]
    list_display_links=['ra_cliente', 'ra']
    date_hierarchy = 'ra_fecha_creacion'
    # raw_id_fields = ('ra_cliente',)
    list_display = [
        'ra',
        'ra_cliente',
        'ra_hora_inicio',
        'ra_hora_final',
        'ra_monto_compra',
        'ra_observacion',
        'ra_fecha_creacion', 
        'ra_user',
        ]
    
    form = Registro_actividadForm

    def DownloadReportSelected(self, request, queryset):
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        wb = Workbook()
        ws=wb.active
        ws['A1'] = 'Reporte por vendedor'+request.GET.urlencode()
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Usuario'
        ws['B2'] = 'suma'
        ws['C2'] = 'Registros'
        ws['D2'] = 'Max Venta'
        ws['E2'] = 'Min Venta'

        queryset = queryset.values('ra_user__username').annotate(
            total=Sum('ra_monto_compra'), 
            conteo=Count('ra'),
            maximo=Max('ra_monto_compra'),
            minimo=Min('ra_monto_compra'),
            ).order_by('ra_user')
        cont = 3

        for item in queryset:
            ws.cell(row=cont, column=1).value = item['ra_user__username']
            ws.cell(row=cont, column=2).value = item['total']
            ws.cell(row=cont, column=3).value = item['conteo']
            ws.cell(row=cont, column=4).value = item['maximo']
            ws.cell(row=cont, column=5).value = item['minimo']
            cont += 1
        

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1


        nombre_archivo='rep_vend.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

        # for it in queryset: 
        #     print(it)
        
    DownloadReportSelected.short_description = "Descargar reporte por vendedor"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.has_perm('plan_de_trabajo.reg_activ_supervisor'):
            qs=qs.all()
        else:
            qs=qs.filter(ra_user=request.user)
        return qs

    def save_model(self, request, obj, form, change):
        obj.ra_user = request.user
        super().save_model(request, obj, form, change)
        

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        if request.user.has_perm('plan_de_trabajo.reg_activ_supervisor'):
            fecha=timezone.localtime(timezone.now()) 
            form.base_fields['ra_hora_inicio'].widget.attrs.update({'value':fecha.time(), 'readonly':True})
            pass
        else:            
            fecha=timezone.localtime(timezone.now()) 
            # fecha_actual=datetime.datetime.now()
            dia_de_la_semana=fecha.weekday()
            str_fecha="{}-{}-{}".format(fecha.year, fecha.month, fecha.day)

            clint=[]
            clint_agregados=[]

            pt_query=Plan_trabajo.objects.filter(pt_vendedor = request.user)
            query_ra=Registro_actividad.objects.filter(ra_user=request.user,ra_fecha_creacion=str_fecha)
            
            for item in pt_query:
                if dia_de_la_semana == 0 and item.pt_lunes:
                    clint.append(item.pt_cliente)
                if dia_de_la_semana == 1 and item.pt_martes:
                    clint.append(item.pt_cliente)
                if dia_de_la_semana == 2 and item.pt_miercoles:
                    clint.append(item.pt_cliente)
                if dia_de_la_semana == 3 and item.pt_jueves:
                    clint.append(item.pt_cliente)
                if dia_de_la_semana == 4 and item.pt_viernes:
                    clint.append(item.pt_cliente)
                if dia_de_la_semana == 5 and item.pt_sabado:
                    clint.append(item.pt_cliente) 
                if dia_de_la_semana == 6 and item.pt_domingo:
                    clint.append(item.pt_cliente)
                

            for item in query_ra:
                clint_agregados.append(item.ra_cliente)
                
            cliente_query=Cliente.objects.filter(cli_clave__in=clint).exclude(cli_clave__in=clint_agregados)
            
            if form.base_fields:
                form.base_fields['ra_cliente'].queryset = cliente_query
            if form.base_fields:
                form.base_fields['ra_hora_inicio'].widget.attrs.update({'value':fecha.time(), 'readonly':True})
        return form
    



class VisitaVendedorConig(admin.ModelAdmin):
    list_filter = [('vv_fecha', DateRangeFilter), 'vv_tipo']       
    search_fields = ['vv_cliente__cli_clave', 'vv_vendedor__username']    
    list_display = [
        'vv_fecha',
        'vv_cliente',
        'vv_vendedor',
        'vv_latitude',
        'vv_longitude',
        'vv_numero_venta',
        'vv_monto_venta', 
        'vv_numero_factura',
        'vv_monto_factura',
        'mapa',
        'vv_tipo',
        ]

admin.site.register(Plan_trabajo, Plan_trabajoAdmin)
admin.site.register(Registro_actividad, AdminRegistroActividad) 
admin.site.register(VisitaVendedor, VisitaVendedorConig) 