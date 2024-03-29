import tempfile
from io import BytesIO

import qrcode
import qrcode.image.svg
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.utils.formats import localize
from django.views.generic import TemplateView, CreateView, ListView, DeleteView, UpdateView, View
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
# Librerias reportlab a usar:
from reportlab.platypus import (SimpleDocTemplate, Spacer,
                                Paragraph, Table, TableStyle)
from svglib.svglib import svg2rlg

from aplicaciones.activos.forms import (CategoriaForm, EspecificacionForm, ActivoForm, AsignacionForm,
                                        AsignacionValidarForm, AsignacionReasignacionForm,
                                        TramiteBajaForm, TramiteBajaValidarForm, TemplateItemForm,
                                        TemplateItemGroupForm, ActivoInitForm)
from aplicaciones.activos.models import (Categoria, Activo, Especificacion, Asignacion, TramiteBaja, ESTADO,
                                         VIDA_ACTIVO, SITUACION, TemplateItem,
                                         TemplateItemGroup)
from aplicaciones.inicio.models import User
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects
from aplicaciones.pedidos.models import Marca

PAGE_WIDTH = letter[0]
PAGE_HEIGHT = letter[1]


class Prueba(TemplateView):
    template_name = 'activos/p.html'


class CategoriaList(ListView):
    model = Categoria
    template_name = 'activos/cat_list.html'

    @method_decorator(permission_required('activos.view_categoria', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CategoriaList, self).dispatch(*args, **kwargs)


class CategoriaCrear(CreateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'activos/categoria_crear.html'
    success_url = reverse_lazy('activos:cat_list')

    @method_decorator(permission_required('activos.add_categoria', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CategoriaCrear, self).dispatch(*args, **kwargs)


class CategoriaUpdate(UpdateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'activos/categoria_crear.html'
    success_url = reverse_lazy('activos:cat_list')

    @method_decorator(permission_required('activos.change_categoria', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CategoriaUpdate, self).dispatch(*args, **kwargs)


class CategoriaDelete(DeleteView):
    model = Categoria
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('activos:cat_list')

    @method_decorator(permission_required('activos.delete_categoria', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(CategoriaDelete, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(CategoriaDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class ActivoList(ListView):
    model = Activo
    template_name = 'activos/activo_list.html'
    paginate_by = 200

    @method_decorator(permission_required('activos.view_activo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ActivoList, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        context['obj_list'] = Categoria.objects.all().order_by('cat_nombre')
        context['list_marca'] = Marca.objects.all().order_by('marca_nombre')
        context['vida'] = VIDA_ACTIVO
        context['situacion'] = SITUACION
        context['total_filter'] = queryset
        context['total_precio'] = queryset.aggregate(Sum('activo_costo'))
        ##############################################################

        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super(ActivoList, self).get_queryset()
        categoria = self.request.GET.get('categoria')
        marca = self.request.GET.get('marca')
        serie = self.request.GET.get('serie')
        barra = self.request.GET.get('barra')
        status = self.request.GET.get('status')
        situacion = self.request.GET.get('situacion')
        if categoria != None:
            queryset = queryset.filter(activo_categoria=categoria)
        if marca != None:
            queryset = queryset.filter(activo_marca=marca)
        if serie != None:
            if serie != '':
                queryset = queryset.filter(activo_serie=serie)
        if barra != None:
            if barra != '':
                queryset = queryset.filter(activo_codigo_barra=barra)
        if status != None:
            queryset = queryset.filter(activo_status=status)
        if situacion != None:
            queryset = queryset.filter(activo_situacion=situacion)

        # FASE DE DESCARGA DE XLS
        # dowload = self.request.GET.get('download')
        # if dowload != None:
        queryset = queryset.order_by('-activo')

        return queryset


class Download_report_activo(TemplateView):
    def get(self, request, *args, **kwargs):
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active

        queryset = Activo.objects.all()
        categoria = self.request.GET.get('categoria')
        marca = self.request.GET.get('marca')
        serie = self.request.GET.get('serie')
        barra = self.request.GET.get('barra')
        status = self.request.GET.get('status')
        situacion = self.request.GET.get('situacion')
        if categoria != None:
            queryset = queryset.filter(activo_categoria=categoria)
        if marca != None:
            queryset = queryset.filter(activo_marca=marca)
        if serie != None:
            if serie != '':
                queryset = queryset.filter(activo_serie=serie)
        if barra != None:
            if barra != '':
                queryset = queryset.filter(activo_codigo_barra=barra)
        if status != None:
            queryset = queryset.filter(activo_status=status)
        if situacion != None:
            queryset = queryset.filter(activo_situacion=situacion)

        ws['A1'] = 'REPORTE DE ACTIVO'
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:I1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '###'
        ws['B2'] = 'Nombre'
        ws['C2'] = 'Categoria'
        ws['D2'] = 'Marca'
        ws['E2'] = 'Modelo'
        ws['F2'] = 'Codigo Barra'
        ws['G2'] = 'Observación'
        ws['H2'] = 'Costo'
        ws['I2'] = 'Asignado'
        ws['J2'] = 'Serie'
        cont = 3

        for item in queryset:
            ws.cell(row=cont, column=1).value = item.activo
            ws.cell(row=cont, column=2).value = item.activo_nombre
            ws.cell(row=cont, column=3).value = item.activo_categoria.cat_nombre
            ws.cell(row=cont, column=4).value = item.activo_marca.marca_nombre
            ws.cell(row=cont, column=5).value = item.activo_modelo
            ws.cell(row=cont, column=6).value = item.activo_codigo_barra
            ws.cell(row=cont, column=7).value = item.activo_observacion
            ws.cell(row=cont, column=8).value = item.activo_costo
            ws.cell(row=cont, column=9).value = str(item.asignado_to())
            ws.cell(row=cont, column=10).value = item.activo_serie
            ws.cell(row=cont, column=8).number_format = '#,##0'
            cont += 1

        penultima_fila = cont - 1
        ws.cell(row=cont, column=7).value = "Total"
        ws["H" + str(cont)] = "=SUM(H3:H{})".format(penultima_fila)
        ws["H" + str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 1

        nombre_archivo = 'ACTIVO_REPORT.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ActivoCrear(CreateView):
    model = Activo
    form_class = ActivoInitForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:activo_list')

    @method_decorator(permission_required('activos.add_activo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ActivoCrear, self).dispatch(*args, **kwargs)

    def form_valid(self, form):
        """If the form is valid, save the associated model."""
        status = form.instance.activo_status
        if status == 5 or status == 6:
            messages.error(self.request, 'No esposible agregar activo en status, (Tramite de baja) o (Baja)',
                           extra_tags='alert-danger')
            context = super().get_context_data()
            return render(self.request, self.template_name, context=context)
        else:
            return super().form_valid(form)


class ActivoUpdate(UpdateView):
    model = Activo
    form_class = ActivoForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:activo_list')

    @method_decorator(permission_required('activos.change_activo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ActivoUpdate, self).dispatch(*args, **kwargs)

    def form_valid(self, form):
        """If the form is valid, save the associated model."""
        status = form.instance.activo_status
        if status == 5 or status == 6:
            messages.error(self.request, 'No esposible agregar activo en status, (Tramite de baja) o (Baja)',
                           extra_tags='alert-danger')
            context = super().get_context_data()
            return render(self.request, self.template_name, context=context)
        else:
            return super().form_valid(form)


class ActivoDelete(DeleteView):
    model = Activo
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('activos:activo_list')

    @method_decorator(permission_required('activos.delete_activo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ActivoDelete, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(ActivoDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class TemplateItemList(ListView):
    model = TemplateItem
    template_name = 'activos/template_item_list.html'


class TemplateItemCreate(CreateView):
    model = TemplateItem
    form_class = TemplateItemForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:tem_item_list')


class TemplateItemUpdate(UpdateView):
    model = TemplateItem
    form_class = TemplateItemForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:tem_item_list')


class TemplateItemDelete(DeleteView):
    model = TemplateItem
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('activos:tem_item_list')

    @method_decorator(permission_required('activos.delete_templateitem', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(TemplateItemDelete, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(TemplateItemDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class TemplateItemGrupoList(ListView):
    model = TemplateItemGroup
    template_name = 'activos/template_itemGroup_list.html'


class TemplateItemGrupoCreate(CreateView):
    model = TemplateItemGroup
    form_class = TemplateItemGroupForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:tem_item_grup_list')


class TemplateItemGrupoUpdate(UpdateView):
    model = TemplateItemGroup
    form_class = TemplateItemGroupForm
    template_name = 'activos/activo_crear.html'
    success_url = reverse_lazy('activos:tem_item_grup_list')


class TemplateItemGrupoDelete(DeleteView):
    model = TemplateItemGroup
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('activos:tem_item_grup_list')

    @method_decorator(permission_required('activos.delete_templateitemgroup', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(TemplateItemGrupoDelete, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(TemplateItemGrupoDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class EspecificacioCreate(CreateView):
    model = Especificacion
    form_class = EspecificacionForm
    template_name = 'activos/especificacion_crear.html'
    success_url = reverse_lazy('activos:cat_list')

    @method_decorator(permission_required('activos.add_especificacion', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(EspecificacioCreate, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['obj_list'] = Especificacion.objects.filter(esp_activo=self.kwargs.get('pk'))
        return context

    def form_valid(self, form):
        form.instance.esp_activo = Activo.objects.get(activo=self.kwargs.get('pk'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('activos:activo_especificacion_crear', kwargs={'pk': self.kwargs.get('pk')})


class EspUpdate(UpdateView):
    model = Especificacion
    form_class = EspecificacionForm
    template_name = 'activos/categoria_crear.html'
    success_url = reverse_lazy('activos:cat_list')

    @method_decorator(permission_required('activos.change_especificacion', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(EspUpdate, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('activos:activo_especificacion_crear', kwargs={'pk': self.kwargs.get('activo')})


class EspDelete(DeleteView):
    model = Especificacion
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('activos:activo_list')

    @method_decorator(permission_required('activos.delete_especificacion', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(EspDelete, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(EspDelete, self).get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        #
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context

    def get_success_url(self):
        return reverse_lazy('activos:activo_especificacion_crear', kwargs={'pk': self.kwargs.get('activo')})


class EspTemplateCreate(TemplateView):
    template_name = "activos/template_atutomatic_crear.html"

    def post(self, request, *args, **kwargs):
        id_item_grop_template = request.POST.get('grup_item_template')
        ID_activo = kwargs.get('id_activo')
        try:
            get_it_grup = TemplateItemGroup.objects.get(itmg=id_item_grop_template)
            for item in get_it_grup.itmg_items.all():
                esp = Especificacion(
                    esp_activo_id=ID_activo,
                    esp_item=item.item_nombre,
                    esp_valor='Sin especificar',
                )
                esp.save()
        except ObjectDoesNotExist:
            pass

        url = reverse('activos:activo_especificacion_crear', kwargs={'pk': ID_activo})
        return redirect(url)

    def get_context_data(self, **kwargs):
        context = super(EspTemplateCreate, self).get_context_data(**kwargs)
        #
        context['TemplateItemGroup'] = TemplateItemGroup.objects.all()

        return context


class AsignacionList(ListView):
    model = Asignacion
    paginate_by = 100
    template_name = 'activos/asignacion_list.html'

    def get_context_data(self, **kwargs):
        # from aplicaciones.pedidos.models import Detalle_pedido, Tipo_Pedido, Pedido
        context = super(AsignacionList, self).get_context_data(**kwargs)
        context['estado'] = ESTADO
        context['usuaris'] = User.objects.all().order_by('username')
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        search_status = self.request.GET.get('search_status')
        search_id = self.request.GET.get('search_id')
        activo_id = self.request.GET.get('activo_id')
        id_user = self.request.GET.get('id_user')
        if search_status is not None:
            queryset = queryset.filter(asig_estado=search_status)
        if search_id is not None:
            queryset = queryset.filter(id=search_id)
        if activo_id is not None:
            queryset = queryset.filter(asig_activo=activo_id)
        if id_user is not None:
            queryset = queryset.filter(asig_user=id_user)

        queryset = queryset.order_by('-id')

        return queryset


class AsignacionCrear(CreateView):
    model = Asignacion
    template_name = "activos/activo_crear.html"
    form_class = AsignacionForm
    success_url = reverse_lazy('activos:activo_asignar_list')

    def form_valid(self, form):
        activo_id = form['asig_activo'].value()
        Activo.objects.filter(activo=activo_id).update(activo_situacion=1)
        form.instance.asig_user_edit = self.request.user
        return super().form_valid(form)


class AsignacionValidar(UpdateView):
    model = Asignacion
    template_name = "activos/activo_crear.html"
    form_class = AsignacionValidarForm
    success_url = reverse_lazy('activos:activo_asignar_list')

    def form_valid(self, form):
        """If the form is valid, save the associated model."""
        form.instance.asig_estado = 1
        ID_activo = form.instance.asig_activo.activo
        Activo.objects.filter(activo=ID_activo).update(activo_situacion=1)
        # self.object = form.save()
        return super().form_valid(form)

    def get_success_url(self):
        """Return the URL to redirect to after processing a valid form."""
        url = "{}?{}".format(self.success_url, self.request.GET.urlencode())

        return url  # success_url may be lazy


class ReasignacionActivo(UpdateView):
    model = Asignacion
    template_name = "activos/activo_crear.html"
    form_class = AsignacionReasignacionForm
    success_url = reverse_lazy('activos:activo_asignar_list')

    def form_valid(self, form):
        """If the form is valid, save the associated model."""
        user_reasignado = form.instance.asig_user
        activo = form.instance.asig_activo
        observacion = form.instance.asig_observacion
        id_asig = self.kwargs.get('pk')
        # ACTUALIZACION DE HISTORICO
        Asignacion.objects.filter(id=id_asig).update(asig_estado=2, asig_user_edit=user_reasignado)
        # CREACION NUEVA ASIGNACION
        new_asig = Asignacion(asig_activo=activo, asig_user=user_reasignado, asig_observacion=observacion,
                              asig_user_edit=self.request.user)
        new_asig.save()

        url = reverse('activos:activo_asignar_list')
        return redirect(url)


class GeneraPdfAsignacion(View):

    def myFirstPage(self, canvas, doc):
        # CABECERA DE PAGINA 
        Title = "RESGUARDO Y ASIGNACIÓN DE EQUIPOS"
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)

        canvas.drawCentredString(PAGE_WIDTH / 2.0, PAGE_HEIGHT - 50, Title)

        # Footer.
        canvas.setFont('Times-Roman', 12)
        canvas.drawString(PAGE_WIDTH / 6, 1.5 * cm, '   Sra. Tere Vázquez Arce')
        canvas.drawString(PAGE_WIDTH / 6, 0.9 * cm,
                          '______________________    __________________    ____________________')
        canvas.drawString(PAGE_WIDTH / 6, 0.5 * cm,
                          'Autoriza Direccion General   Asigna Sub Gerente TI    Nombre y Firma usuario')

        canvas.restoreState()

    def get_espfic(self, list, stylo):
        temporal = []
        for item in list:
            temporal.append(Paragraph("✔ {}".format(item.esp_item), stylo))
        return temporal

    def make_qr_code_drawing(self, data, size):
        qr = qrcode.QRCode(version=2, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=size, border=4)
        qr.add_data(data)
        qrcode_svg = qr.make_image(image_factory=qrcode.image.svg.SvgPathFillImage)
        svg_file = tempfile.NamedTemporaryFile()
        qrcode_svg.save(svg_file)  # store as an SVG file
        svg_file.flush()
        qrcode_rl = svg2rlg(svg_file.name)  # load SVG file as reportlab graphics
        svg_file.close()
        return qrcode_rl

    def get(self, request, *args, **kwargs):
        id_resguardo = self.kwargs.get('id_resguardo')
        get_asignacion = Asignacion.objects.get(id=id_resguardo)

        response = HttpResponse(content_type='application/pdf')
        pdf_name = "clientes.pdf"  # llamado clientes
        # la linea 26 es por si deseas descargar el pdf a tu computadora
        # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
        buff = BytesIO()

        doc = SimpleDocTemplate(buff,
                                pagesize=letter,
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                title='Asignacion de activo'
                                )

        folio_format = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=10, fontName="Times-Roman")
        # title = ParagraphStyle('titulo', alignment = TA_CENTER, fontSize = 12, fontName="Times-Roman")
        parrafo2 = ParagraphStyle('parrafo', alignment=TA_RIGHT, fontSize=10, fontName="Times-Roman")
        parrafo = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=10, fontName="Times-Roman")
        aviso_empresa_style = ParagraphStyle('parrafo', alignment=TA_JUSTIFY, fontSize=8, fontName="Times-Roman",
                                             spaceBefore=15)

        tabla_body = ParagraphStyle('t_body', alignment=TA_LEFT, fontSize=8, fontName="Times-Roman")
        tabla_head = ParagraphStyle('t_head', alignment=TA_CENTER, fontSize=9, fontName="Times-Roman")

        items = []

        folio_txt = "<strong>FOLIO N°:</strong><em><u>{}</u></em>".format(get_asignacion.id)
        p = Paragraph(folio_txt, folio_format)
        items.append(p)

        txt2 = "<strong>Fecha:</strong>{}".format(localize(get_asignacion.asig_fecha_adicion))
        p = Paragraph(txt2, parrafo2)
        items.append(p)

        texto = "<strong>Usuario:</strong> <em><u>{}</u></em>  <strong>Departamento:</strong><em><u>{}</u></em>".format(
            get_asignacion.asig_user.get_full_name(), get_asignacion.asig_user.departamento)
        p = Paragraph(texto, parrafo)
        items.append(p)

        texto = "Por este medio Comercializadora computel del sureste me asigna las siguientes claves, derechos y equipos para desarrollar las actividades para que fui contratado, es mi responsabilidad, el uso, manejo y protección de la información que en ellos pose, asi como mantener limpios lo equipos y tener los debidos cuidados para mantenerse en buen estado."
        p = Paragraph(texto, aviso_empresa_style)
        items.append(p)

        texto = "De detectar cambios en el sofware/hardware(programas no autorizados) seré sancionado de acuerdo a la gravedad y segun las politicas del departamento de TI."
        p = Paragraph(texto, aviso_empresa_style)
        items.append(p)

        t_titulos = (
            Paragraph('EQUIPO', tabla_head),
            Paragraph('DESCRIPCIÓN', tabla_head),
            Paragraph('ADICIONAL', tabla_head),
            Paragraph('SERIE', tabla_head),
            Paragraph('COD. INVENT', tabla_head),
            Paragraph('MARCA/MODELO', tabla_head))
        t_body = [(
            Paragraph(get_asignacion.asig_activo.activo_categoria.cat_nombre, tabla_body),
            Paragraph(get_asignacion.asig_activo.activo_nombre, tabla_body),
            self.get_espfic(Especificacion.objects.filter(esp_activo=get_asignacion.asig_activo), tabla_body),
            Paragraph(get_asignacion.asig_activo.activo_serie, tabla_body),
            Paragraph(str(get_asignacion.asig_activo.activo), tabla_body),
            Paragraph("{}/{}".format(get_asignacion.asig_activo.activo_marca, get_asignacion.asig_activo.activo_modelo),
                      tabla_body)
        )]
        t = Table([t_titulos] + t_body, colWidths=[2 * cm, 3 * cm, 4 * cm, 3 * cm, 2 * cm, 4 * cm])
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (5, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))

        items.append(t)

        texto = "<strong>OBSERVACIÓN:</strong>{}".format(get_asignacion.asig_observacion)
        p = Paragraph(texto, aviso_empresa_style)
        items.append(p)

        current_site = str(get_current_site(self.request))
        revers = reverse_lazy('activos:activo_asignar_pdf', args={id_resguardo: get_asignacion.id})

        url = "http://{}{}".format(current_site, revers)

        qrcode_rl = self.make_qr_code_drawing(url, 8)

        items.append(qrcode_rl)

        doc.build(items, onFirstPage=self.myFirstPage)
        response.write(buff.getvalue())
        buff.close()
        return response


class TramiteBajaList(ListView):
    model = TramiteBaja
    paginate_by = 100
    template_name = 'activos/tramite_baja_list.html'
    ordering = ['-tb_activo']

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.order_by('-id')
        return queryset

    def get_context_data(self, **kwargs):
        # from aplicaciones.pedidos.models import Detalle_pedido, Tipo_Pedido, Pedido
        context = super().get_context_data(**kwargs)
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context


class PDFBajaView(View):

    def myFirstPage(self, canvas, doc):
        # CABECERA DE PAGINA 
        Title = "BAJA DE ACTIVO"
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)

        canvas.drawCentredString(PAGE_WIDTH / 2.0, PAGE_HEIGHT - 50, Title)

        # Footer.
        canvas.setFont('Times-Roman', 8)
        # canvas.drawString( PAGE_WIDTH/6, 1.5*cm, '   Sra. Tere Vázquez Arce')
        canvas.drawString(PAGE_WIDTH / 6, 1 * cm,
                          '_______________________________            __________________________________            ____________________')
        canvas.drawString(PAGE_WIDTH / 6, 0.7 * cm,
                          '   Ing. Victor Manuel Galan Robles                           C. Teresa de Jesús Vázquez Arce                     Encargado de activos')
        canvas.drawString(PAGE_WIDTH / 6, 0.4 * cm,
                          '                  Subgerente de TI                                          Dirección general adjunta')

        canvas.restoreState()

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/pdf')
        buff = BytesIO()
        doc = SimpleDocTemplate(buff,
                                pagesize=letter,
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                title='Baja de activo'
                                )

        baja_ID = request.GET.get('item')
        baja_obj = TramiteBaja.objects.get(id=baja_ID)

        folio_format = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=10, fontName="Times-Roman")
        # title = ParagraphStyle('titulo', alignment = TA_CENTER, fontSize = 12, fontName="Times-Roman")
        parrafo2 = ParagraphStyle('parrafo', alignment=TA_RIGHT, fontSize=10, fontName="Times-Roman")
        parrafo = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=10, fontName="Times-Roman")
        aviso_empresa_style = ParagraphStyle('parrafo', alignment=TA_JUSTIFY, fontSize=8, fontName="Times-Roman",
                                             spaceBefore=15)

        tabla_body = ParagraphStyle('t_body', alignment=TA_LEFT, fontSize=8, fontName="Times-Roman")
        tabla_head = ParagraphStyle('t_head', alignment=TA_CENTER, fontSize=9, fontName="Times-Roman")

        TITULO = ParagraphStyle('TITULO', alignment=TA_CENTER, fontSize=14, fontName="Times-Roman")
        BODY = ParagraphStyle('BODY', alignment=TA_LEFT, fontSize=11, fontName="Times-Roman")

        items = []

        folio_txt = "<strong>FOLIO N°:</strong><em><u>{}</u></em>".format(baja_obj.id)
        p = Paragraph(folio_txt, folio_format)
        items.append(p)

        folio_txt = "Villahermosa, Tabasco, a {}".format(localize(baja_obj.tb_fecha_creacion))
        p = Paragraph(folio_txt, parrafo2)
        items.append(p)

        folio_txt = "<strong>BAJA</strong>"
        p = Paragraph(folio_txt, TITULO)
        items.append(p)

        items.append(Spacer(1, 0.2 * cm))
        folio_txt = "<strong>Usuario validó: </strong> {} <strong>Observación: </strong> {}".format(
            baja_obj.tb_user_valido, baja_obj.tb_observacion)
        p = Paragraph(folio_txt, BODY)
        items.append(p)

        items.append(Spacer(1, 0.2 * cm))
        folio_txt = "<strong>ASIGNACION</strong>"
        p = Paragraph(folio_txt, TITULO)
        items.append(p)

        items.append(Spacer(1, 0.2 * cm))
        folio_txt = "<strong>USUARIO</strong> {} <strong>OBSERVACION:</strong>{}".format(
            baja_obj.tb_activo.asig_user.get_full_name(),
            baja_obj.tb_activo.asig_observacion,
        )
        p = Paragraph(folio_txt, BODY)
        items.append(p)

        items.append(Spacer(1, 0.2 * cm))
        folio_txt = "<strong>DESCRIPCION DE EQUIPO</strong>"
        p = Paragraph(folio_txt, TITULO)
        items.append(p)

        t_titulos = (
            Paragraph('INFORMACION', tabla_head),
            Paragraph('DESCRIPCION', tabla_head),
        )
        t_body = [
            (Paragraph("ACTIVO", tabla_body), Paragraph(baja_obj.tb_activo.asig_activo.activo_nombre, tabla_body)),
            (Paragraph("CATEGORIA", tabla_body),
             Paragraph(str(baja_obj.tb_activo.asig_activo.activo_categoria), tabla_body)),
            (Paragraph("MARCA", tabla_body), Paragraph(str(baja_obj.tb_activo.asig_activo.activo_marca), tabla_body)),
            (Paragraph("MODELO", tabla_body), Paragraph(str(baja_obj.tb_activo.asig_activo.activo_modelo), tabla_body)),
            (Paragraph("SERIE", tabla_body), Paragraph(baja_obj.tb_activo.asig_activo.activo_serie, tabla_body)),
            (
            Paragraph("CODIGO", tabla_body), Paragraph(baja_obj.tb_activo.asig_activo.activo_codigo_barra, tabla_body)),
            (Paragraph("STATUS", tabla_body),
             Paragraph(baja_obj.tb_activo.asig_activo.get_activo_status_display(), tabla_body)),
            (Paragraph("SITUACION", tabla_body),
             Paragraph(baja_obj.tb_activo.asig_activo.get_activo_situacion_display(), tabla_body)),
            (Paragraph("OBSERVACION", tabla_body),
             Paragraph(baja_obj.tb_activo.asig_activo.activo_observacion, tabla_body)),
        ]
        t = Table([t_titulos] + t_body, colWidths=[3 * cm, 10 * cm])

        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (5, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))

        items.append(Spacer(1, 0.5 * cm))
        items.append(t)

        doc.build(items, onFirstPage=self.myFirstPage)
        response.write(buff.getvalue())
        buff.close()
        return response


class TramiteBajaCrear(CreateView):
    model = TramiteBaja
    template_name = "activos/activo_crear.html"
    form_class = TramiteBajaForm
    success_url = reverse_lazy('activos:tb_list')

    def form_valid(self, form):
        """If the form is valid, save the associated model."""
        form.instance.tb_user_edit = self.request.user
        asignacion = form.instance.tb_activo
        asg = Asignacion.objects.get(id=asignacion.id)
        Activo.objects.filter(activo=asg.asig_activo.activo).update(activo_status=5)
        return super().form_valid(form)


class TramiteBajaUpdate(UpdateView):
    model = TramiteBaja
    template_name = "activos/activo_crear.html"
    form_class = TramiteBajaValidarForm
    success_url = reverse_lazy('activos:tb_list')

    def form_valid(self, form):
        """If the form is valid, save the associated model."""

        form.instance.tb_user_valido = self.request.user
        asignacion = form.instance.tb_activo.id
        asg = Asignacion.objects.get(id=asignacion)
        Asignacion.objects.filter(id=asignacion).update(asig_estado=2)
        Activo.objects.filter(activo=asg.asig_activo.activo).update(activo_status=6)
        return super().form_valid(form)


class MiAsignacionList(ListView):
    model = Asignacion
    paginate_by = 100
    template_name = 'activos/asignacion_list.html'

    def get_context_data(self, **kwargs):
        # from aplicaciones.pedidos.models import Detalle_pedido, Tipo_Pedido, Pedido
        context = super(MiAsignacionList, self).get_context_data(**kwargs)
        context['estado'] = ESTADO
        # context['usuaris'] = User.objects.all().order_by('username')
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(asig_user=self.request.user)
        search_status = self.request.GET.get('search_status')
        search_id = self.request.GET.get('search_id')
        activo_id = self.request.GET.get('activo_id')
        id_user = self.request.GET.get('id_user')
        if search_status != None:
            queryset = queryset.filter(asig_estado=search_status)

        if search_id != None:
            queryset = queryset.filter(id=search_id)
        if activo_id != None:
            queryset = queryset.filter(asig_activo=activo_id)
        if id_user != None:
            queryset = queryset.filter(asig_user=id_user)

        return queryset
