from django.shortcuts import render
from django.views.generic import TemplateView, ListView, UpdateView, DeleteView, View
from aplicaciones.empresa.models import Cliente
from aplicaciones.expo.models import AsignacionMarca, VentaExpo, Detalle_venta, TIPO_VENTA, AsignacionVendedor_a_Supervisor
from aplicaciones.pedidos.models import Producto
from django.db.models import Q, Sum, F, FloatField, Count
from django.shortcuts import redirect, reverse
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.urls import reverse_lazy

from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects

from aplicaciones.expo.forms import VentaExpoForm, EditProductoExpoForm
# Create your views here.

# import para generar xls
from openpyxl.styles import Font, Fill, Alignment, Color, PatternFill
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from decimal import Decimal

from django.utils.formats import localize

# Librerias reportlab a usar:
from django.http import HttpResponse
from io import BytesIO
from django.conf import settings
from reportlab.platypus import (SimpleDocTemplate, PageBreak, Image, Spacer,
                                Paragraph, Table, TableStyle, Spacer, BaseDocTemplate, Frame, PageTemplate)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_RIGHT, TA_LEFT
PAGE_WIDTH = letter[0]
PAGE_HEIGHT = letter[1]


from aplicaciones.pedidos.models import Marca


from django.utils.formats import localize
from django.contrib.humanize.templatetags.humanize import intcomma
from datetime import datetime
class SelectCienteView(TemplateView):
    template_name = "expo/select_cliente.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        Buscar = self.request.GET.get('Buscar')
        if Buscar != None:
            context['cliente_list'] = Cliente.objects.filter(
                Q(cli_clave=Buscar) | Q(cli_nombre__icontains=Buscar))
        return context


class VentaView(TemplateView):
    template_name = "expo/ventalist.html"
    tipo_venta = 0
    no_venta = 0

    def dispatch(self, *args, **kwargs):
        cliente = self.request.GET.get('cleint')
        tipo_vent = self.request.GET.get('tip')

        if self.request.GET.get('end') != None:
            VentaExpo.objects.filter(Venta_ID=self.request.GET.get(
                'venta')).update(venta_e_status=True)
            url = reverse('expo:selec_cliente')
            return redirect(url)

        if self.request.GET.get('venta') == None:
            obj = VentaExpo(
                venta_e_cliente_id=cliente,
                venta_e_creado=self.request.user,
                venta_e_tipo=tipo_vent,
            )
            obj.save()
            url = reverse('expo:venta')
            url = "{}?cleint={}&tip={}&venta={}".format(
                url, cliente, tipo_vent, obj.Venta_ID)
            return redirect(url)
        return super(VentaView, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        venta = self.request.GET.get('venta')
        tip = self.request.GET.get('tip')
        tipo_venta_display=""
        for item in TIPO_VENTA:
            if str(item[0]) == tip:
                tipo_venta_display = item[1]
        try:
            asignacion = AsignacionMarca.objects.get(am_user=self.request.user)
            filtro = asignacion.am_marca.values_list('marca_id_marca')
        except ObjectDoesNotExist:
            filtro = []
            messages.warning(self.request, 'Es necesario que se le asigne una marca al usuario "' +
                             str(self.request.user) + '" posteriormente actualice la pagina')

        context['productos_list'] = Producto.objects.filter(producto_marca__in=filtro, tipo_producto=tip).order_by('producto_descripcion')
        sum_detalle = Detalle_venta.objects.filter(detalle_venta=venta).aggregate(total=Sum(F('detalle_cantidad')*F('detalle_precio'), output_field=FloatField()))['total']
        context['total_venta'] = round(sum_detalle, 3) if sum_detalle != None else sum_detalle
        context['tipo_venta_display'] = tipo_venta_display

        return context

    def post(self, request, *args, **kwargs):
        from django.http import JsonResponse
        cliente = self.request.POST.get('cleint')
        tipo_vent = self.request.POST.get('tip')

        venta = self.request.POST.get('no_venta')
        cantidad = self.request.POST.get('cantidad')
        producto = self.request.POST.get('producto')

        # PROCESO DE VENTA
        obj_producto = Producto.objects.get(producto_codigo=producto)

        obj_detalle_venta = Detalle_venta(
            detalle_venta_id=venta,
            detalle_producto_id=obj_producto,
            detalle_cantidad=cantidad,
            detalle_precio=obj_producto.producto_precio,
        )
        obj_detalle_venta.save()

        sum_detalle = Detalle_venta.objects.filter(detalle_venta=venta).aggregate(total=Sum(F('detalle_cantidad')*F('detalle_precio'), output_field=FloatField()))['total']
        total_venta = round(sum_detalle, 3) if sum_detalle != None else sum_detalle
        data = {
            'producto': producto,
            'total_venta': total_venta,
        }
        return JsonResponse(data)
        


class VentaList(ListView):
    template_name = "expo/ventas.html"
    model = VentaExpo
    paginate_by = 200

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sum_detalle=0;
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada 
        context['Clientes'] = VentaExpo.objects.values('venta_e_cliente__cli_clave').annotate(dcount=Count('venta_e_cliente__cli_clave')).order_by('venta_e_cliente__cli_clave')
        context['creadors'] = VentaExpo.objects.values('venta_e_creado__id', 'venta_e_creado__username').annotate(dcount=Count('venta_e_creado')).order_by('venta_e_creado')
        context['marca'] = Marca.objects.all()
        context['tipo_venta'] = TIPO_VENTA

        
        

        return context

    def get_queryset(self):
        from django.contrib.humanize.templatetags.humanize import intcomma
        queryset = super().get_queryset()
        sum_detalle=0
        if self.request.user.is_superuser == False:
            try:
                supervisor = AsignacionVendedor_a_Supervisor.objects.get(
                    avs_Supervisor=self.request.user)
                vendedores = supervisor.avs_vendedors.values_list('id')
                queryset = queryset.filter(
                    venta_e_creado__in=vendedores).order_by('venta_e_fecha_pedido')
            except ObjectDoesNotExist:
                queryset = queryset.none()
                messages.warning(self.request, 'Es necesario que se le asigne vendedores al usuario "' + str(
                    self.request.user) + '" posteriormente actualice la pagina')

        venta_e_fecha_pedido_init = self.request.GET.get('venta_e_fecha_pedido_init')
        venta_e_fecha_pedido_end = self.request.GET.get('venta_e_fecha_pedido_end')
        venta_e_cliente = self.request.GET.get('venta_e_cliente')
        venta_e_status = self.request.GET.get('venta_e_status')
        venta_e_tipo = self.request.GET.get('venta_e_tipo')
        venta_e_creado = self.request.GET.get('venta_e_creado')
        Buscar = self.request.GET.get('Buscar')

        if venta_e_fecha_pedido_init != None and venta_e_fecha_pedido_init != "":
            if venta_e_fecha_pedido_end != None and venta_e_fecha_pedido_end != "":
                queryset = queryset.filter(venta_e_fecha_pedido__range=( venta_e_fecha_pedido_init, venta_e_fecha_pedido_end))
        if venta_e_cliente != None and venta_e_cliente != "":
            queryset = queryset.filter(venta_e_cliente__cli_clave=venta_e_cliente)
        if venta_e_status != None and venta_e_status != "":
            queryset = queryset.filter(venta_e_status=venta_e_status)
        if venta_e_tipo != None and venta_e_tipo != "":
            queryset = queryset.filter(venta_e_tipo=venta_e_tipo)
        if Buscar != None and Buscar != "":
            queryset = queryset.filter(Venta_ID=Buscar)
        if venta_e_creado != None and venta_e_creado != "":
            queryset = queryset.filter(venta_e_creado=venta_e_creado)

        
        for item in queryset:
            sum_detalle = sum_detalle + item.total_venta()
        count_row_ventas=queryset.count()
        suma= round(sum_detalle, 3) if sum_detalle != None else sum_detalle
        messages.info(self.request, "Total ventas: ${}. N° de ventas:{}".format(intcomma(suma), count_row_ventas))

        

        return queryset


class dowload_ventas_expo(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        queryset = VentaExpo.objects.all()

        if self.request.user.is_superuser == False:
            try:
                supervisor = AsignacionVendedor_a_Supervisor.objects.get(
                    avs_Supervisor=self.request.user)
                vendedores = supervisor.avs_vendedors.values_list('id')
                queryset = queryset.filter(
                    venta_e_creado__in=vendedores).order_by('venta_e_fecha_pedido')
            except ObjectDoesNotExist:
                queryset = queryset.none()
                messages.warning(self.request, 'Es necesario que se le asigne vendedores al usuario "' + str(
                    self.request.user) + '" posteriormente actualice la pagina')

        venta_e_fecha_pedido_init = self.request.GET.get('venta_e_fecha_pedido_init')
        venta_e_fecha_pedido_end = self.request.GET.get('venta_e_fecha_pedido_end')
        venta_e_cliente = self.request.GET.get('venta_e_cliente')
        venta_e_status = self.request.GET.get('venta_e_status')
        venta_e_tipo = self.request.GET.get('venta_e_tipo')
        venta_e_creado = self.request.GET.get('venta_e_creado')
        Buscar = self.request.GET.get('Buscar')

        if venta_e_fecha_pedido_init != None and venta_e_fecha_pedido_init != "":
            if venta_e_fecha_pedido_end != None and venta_e_fecha_pedido_end != "":
                queryset = queryset.filter(venta_e_fecha_pedido__range=( venta_e_fecha_pedido_init, venta_e_fecha_pedido_end))
        if venta_e_cliente != None and venta_e_cliente != "":
            queryset = queryset.filter(venta_e_cliente__cli_clave=venta_e_cliente)
        if venta_e_status != None and venta_e_status != "":
            queryset = queryset.filter(venta_e_status=venta_e_status)
        if venta_e_tipo != None and venta_e_tipo != "":
            queryset = queryset.filter(venta_e_tipo=venta_e_tipo)
        if Buscar != None and Buscar != "":
            queryset = queryset.filter(Venta_ID=Buscar)
        if venta_e_creado != None and venta_e_creado != "":
            queryset = queryset.filter(venta_e_creado=venta_e_creado)

        ws['A1'] = 'Ventas Expo'
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:G1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '##'
        ws['B2'] = 'Fecha'
        ws['C2'] = 'Actualizado'
        ws['D2'] = 'Cliente'
        ws['E2'] = 'Creador'
        ws['F2'] = 'Tipo'
        ws['G2'] = 'Total'
        ws['H2'] = 'Status'
        cont = 3
  

        for item in queryset:
            ws.cell(row=cont, column=1).value = item.Venta_ID
            ws.cell(row=cont, column=2).value = localize(item.venta_e_fecha_pedido)
            ws.cell(row=cont, column=3).value = localize(item.venta_e_actualizado)
            ws.cell(row=cont, column=4).value = item.venta_e_cliente.cli_clave
            ws.cell(row=cont, column=5).value = str(item.venta_e_creado)
            ws.cell(row=cont, column=6).value = item.get_venta_e_tipo_display()
            ws.cell(row=cont, column=7).value = item.total_venta()
            ws.cell(row=cont, column=8).value = item.get_status_display()
            # ws.cell(row=cont, column=8).fill = PatternFill(bgColor="00ff7f", fill_type = "solid")
            ws.cell(row=cont, column=7).number_format = '#,##0'
            cont += 1

        ws["F"+str(cont)] = "TOTAL"
        ws["G"+str(cont)] = "=SUM(G3:G"+str(cont-1)+")"
        ws["G"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'VENTAS_EXPO.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class VentaDelete(DeleteView):
    model = VentaExpo
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('expo:venta_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context

    @method_decorator(permission_required('expo.delete_ventaexpo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class VentaExpoUpdate(UpdateView):
    model = VentaExpo
    template_name = "expo/formulario.html"
    form_class = VentaExpoForm
    success_url = reverse_lazy('expo:venta_list')

    @method_decorator(permission_required('expo.change_ventaexpo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class DetalleVentaList(ListView):
    template_name = "expo/detalleventas.html"
    model = Detalle_venta
    paginate_by = 200

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        urls_formateada = self.request.GET.copy()
        producto = self.request.GET.get('find_prod')
        if 'page' in urls_formateada:
            del urls_formateada['page']
    
        context['productos_list'] = Producto.objects.filter(producto_codigo=producto)
        
        
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        venta_id = self.kwargs['pk']
        queryset = queryset.filter(detalle_venta=venta_id)
        return queryset

    
    def post(self, request, *args, **kwargs):

        venta = self.request.POST.get('no_venta')
        cantidad = self.request.POST.get('cantidad')
        producto = self.request.POST.get('producto')

        # PROCESO DE VENTA
        producto_obj=Producto.objects.get(producto_codigo=producto)
        venta_expo_obj = VentaExpo.objects.get(Venta_ID=venta)

        detalle_venta=Detalle_venta(
            detalle_venta = venta_expo_obj,
            detalle_producto_id = producto_obj,
            detalle_cantidad = cantidad,
            detalle_precio = producto_obj.producto_precio,
            )
        print(producto_obj.producto_precio)
        detalle_venta.save()
        url = reverse_lazy('expo:detalle_vent', kwargs = {'pk':venta})
        return redirect(url)



class DetalleVentaDelete(DeleteView):
    model = Detalle_venta
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('expo:detalle_vent')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context

    def get_success_url(self):
        vent = self.request.GET.get('vent')
        url = reverse_lazy('expo:detalle_vent', kwargs = {'pk':vent})
        return url



    @method_decorator(permission_required('expo.delete_detalle_venta', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class ProductoListProveedor(ListView):
    model = Producto
    ordering = ['producto_descripcion']
    template_name = "pedidos/producto/list_producto.html"
    paginate_by = 100

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        producto_categoria = self.request.GET.getlist('producto_categoria')
        tipo_producto = self.request.GET.getlist('tipo_producto')
        checkbox_visible = self.request.GET.getlist('checkbox_visible')
        Buscar = self.request.GET.get('Buscar')
        try:
            asignacion = AsignacionMarca.objects.get(am_user=self.request.user)
            filtro = asignacion.am_marca.values_list('marca_id_marca')
            queryset = queryset.filter(producto_marca__in=filtro)
            if len(tipo_producto) != 0:
                queryset = queryset.filter(tipo_producto__in=tipo_producto)
            if len(producto_categoria) != 0:
                queryset = queryset.filter(
                    producto_categoria__in=producto_categoria)
            if len(checkbox_visible) != 0:
                queryset = queryset.filter(
                    producto_visible__in=checkbox_visible)
            if Buscar != None:
                queryset = queryset.filter(Q(producto_codigo=Buscar) | Q(
                    producto_nombre__icontains=Buscar))
        except ObjectDoesNotExist:
            queryset = queryset.none()
            messages.warning(self.request, 'Es necesario que se le asigne una marca al usuario "' + \
                             str(self.request.user) + '" posteriormente actualice la pagina')
        return queryset

    @method_decorator(permission_required('expo.puede_ver_producto_expo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class ProductoExpoEdit(UpdateView):
    model = Producto
    form_class = EditProductoExpoForm
    template_name = "expo/form_producto.html"
    success_url = reverse_lazy('expo:producto_list_proveedor')

    @method_decorator(permission_required('expo.puede_editar_producto_expo', reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)




class dowload_venta_expo_ID(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        VentaID = request.GET.get('vent')
        queryset = Detalle_venta.objects.filter(detalle_venta=VentaID)
        

        ws['A1'] = 'Detalle de venta N° {}'.format(VentaID)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Cliente'
        ws['B2'] = 'Producto'
        ws['C2'] = 'Descripción'
        ws['D2'] = 'Cantidad'
        ws['E2'] = 'Precio'
        ws['F2'] = 'Subtotal'
        cont = 3

        for item in queryset:
            ws.cell(row=cont, column=1).value = item.detalle_venta.venta_e_cliente.cli_clave
            ws.cell(row=cont, column=2).value = item.detalle_producto_id.producto_codigo
            ws.cell(row=cont, column=3).value = item.detalle_producto_id.producto_descripcion
            ws.cell(row=cont, column=4).value = item.detalle_cantidad
            ws.cell(row=cont, column=5).value = item.detalle_precio
            ws.cell(row=cont, column=6).value = item.subtotal()
            ws.cell(row=cont, column=6).number_format = '#,##0'
            cont += 1

        ws["E"+str(cont)] = "TOTAL"
        ws["F{}".format(cont)] = "=SUM(F3:F{})".format(cont-1)
        ws["F"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'venta_no_{}.xls'.format(VentaID)
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



class Pdf_recibo_cliente(View):
    clave_cliente = None
    venta_e_status = None

    def dispatch(self, *args, **kwargs):
        self.clave_cliente = self.request.GET.get('venta_e_cliente')
        self.venta_e_status = self.request.GET.get('venta_e_status')
        return super().dispatch(*args, **kwargs)

    def myFirstPage(self, canvas, doc):
        status_ventas = ""
        if self.venta_e_status:
            status_ventas =  "Finalizada"
        else:
            status_ventas = "Inconclusa"
        # CABECERA DE PAGINA
        Title = "RECIBO DE CLIENTE N° {}  {}".format(self.clave_cliente, status_ventas)
        canvas.saveState()
        canvas.setFont('Times-Bold', 16)

        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 50, Title)
        # Logo de empresa
        # archivo_imagen = self.object_catalogo.tp_empresa.empresa_logo.path
        # Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
        # canvas.drawImage(archivo_imagen, 20, 690, 120, 90, preserveAspectRatio=True)

        canvas.saveState()
        canvas.setFont('Times-Roman', 10)

        page_count = doc.page
        canvas.drawCentredString(PAGE_WIDTH/2.0,PAGE_HEIGHT - 770, "Página {}".format(page_count))
        canvas.restoreState()

        canvas.setFont('Times-Roman', 10)

        canvas.drawString(420, 30, '{}'.format(localize(datetime.now())))

    def myLaterPages(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Roman', 10)

        page_count = doc.page
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 770, "Página {}".format(page_count))
        canvas.drawString(420, 30, '{}'.format(localize(datetime.now())))
        canvas.restoreState()

    

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/pdf')
        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=30, title='RECIBO CLIENTE {}'.format(self.clave_cliente))
        items = []
        data_tabla = []

        stylo_p_center = ParagraphStyle('parrafo_center', alignment=TA_CENTER, fontSize=11, fontName="Times-Roman")
        stylo_p_center_INFO = ParagraphStyle('parrafo_center', alignment=TA_CENTER, fontSize=8, fontName="Times-Roman")
        stylo_p = ParagraphStyle('parrafo', alignment=TA_LEFT, fontSize=11, fontName="Times-Roman")
        stylo_titulo = ParagraphStyle('titulo', alignment=TA_CENTER, fontSize=11, fontName="Times-Bold")
        stylo_portada_title = ParagraphStyle('titulo', alignment=TA_CENTER, fontSize=20, fontName="Times-Bold")

        
        # items.append(Spacer(0, 30))
        # items.append(PageBreak())
        marcas = []
        detale_venta_expo_obj = Detalle_venta.objects.filter(
            detalle_venta__venta_e_cliente__cli_clave=self.clave_cliente, 
            detalle_venta__venta_e_status=self.venta_e_status
            ).order_by('detalle_venta__Venta_ID', 'detalle_producto_id__producto_marca')
        dta = []
        
        titulos_tabla = [(
            Paragraph('N°V', stylo_titulo), 
            Paragraph('Producto', stylo_titulo), 
            Paragraph('Marca', stylo_titulo), 
            Paragraph('Descripción', stylo_titulo),
            Paragraph('Cantidad', stylo_titulo),
            Paragraph('Precio', stylo_titulo),
            Paragraph('Importe', stylo_titulo),
            )]


        
        
        for item in detale_venta_expo_obj:                   
            dta.append((
                Paragraph(str(item.detalle_venta.Venta_ID), stylo_p_center), 
                Paragraph(item.detalle_producto_id.producto_codigo, stylo_p_center), 
                Paragraph(item.detalle_producto_id.producto_marca.marca_nombre, stylo_p_center), 
                Paragraph(item.detalle_producto_id.producto_descripcion, stylo_p),
                Paragraph(str(item.detalle_cantidad), stylo_p_center), 
                Paragraph(intcomma(item.detalle_precio), stylo_p_center), 
                Paragraph(intcomma(item.subtotal()), stylo_p_center), 
                ))
            
           

        
        sum_detalle_marca = Detalle_venta.objects.values(
            'detalle_producto_id__producto_marca',
            'detalle_producto_id__producto_marca__marca_nombre'
        ).filter( 
            detalle_venta__venta_e_cliente__cli_clave=self.clave_cliente, 
            detalle_venta__venta_e_status=self.venta_e_status
            ).annotate(total=Sum(F('detalle_cantidad')*F('detalle_precio'), output_field=FloatField())).order_by('detalle_producto_id__producto_marca')
        


        for detalle_marca in sum_detalle_marca:
            dta.append((
                "",
                "",
                "",
                "",
                Paragraph(detalle_marca['detalle_producto_id__producto_marca__marca_nombre'], stylo_p_center_INFO),
                Paragraph("<-- subtotal", stylo_p_center_INFO),
                Paragraph(intcomma(detalle_marca['total']), stylo_p_center_INFO),
            ))




        sum_detalle = Detalle_venta.objects.filter( 
            detalle_venta__venta_e_cliente__cli_clave=self.clave_cliente, 
            detalle_venta__venta_e_status=self.venta_e_status
            ).aggregate(total=Sum(F('detalle_cantidad')*F('detalle_precio'), output_field=FloatField()))['total']
        total_venta = round(sum_detalle, 3) if sum_detalle != None else sum_detalle

        dta.append((
                "",
                "",
                "",
                "",
                "",
                Paragraph("TOTAL", stylo_p_center),
                Paragraph(intcomma(total_venta), stylo_p_center),
            ))

            
        

            

        tabla = Table(titulos_tabla+dta, colWidths=[
            1.5 * cm, 
            2.5 * cm,
            2.5 * cm, 
            7.5 * cm, 
            2 * cm,
            2 * cm,
            2 * cm,
            ])

        tabla.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.dodgerblue),
                # ('LINEBELOW', (0, 0), (-1, 0), 0, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.transparent)
            ]
        ))
        items.append(tabla)
        items.append(Spacer(0, 20))

        doc.build(items, onFirstPage=self.myFirstPage,onLaterPages=self.myLaterPages)
        response.write(buff.getvalue())
        buff.close()
        return response









class TotalVentasMarcasDate(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        queryset = Detalle_venta.objects.all()

        venta_e_fecha_pedido_init = self.request.GET.get('venta_e_fecha_pedido_init')
        venta_e_fecha_pedido_end = self.request.GET.get('venta_e_fecha_pedido_end')
        venta_e_cliente = self.request.GET.get('venta_e_cliente')
        venta_e_status = self.request.GET.get('venta_e_status')
        venta_e_tipo = self.request.GET.get('venta_e_tipo')
        venta_e_creado = self.request.GET.get('venta_e_creado')
        
        

        if venta_e_fecha_pedido_init != None and venta_e_fecha_pedido_init != "":
            if venta_e_fecha_pedido_end != None and venta_e_fecha_pedido_end != "":
                queryset = queryset.filter(detalle_venta__venta_e_fecha_pedido__range=[venta_e_fecha_pedido_init, venta_e_fecha_pedido_end])
        if venta_e_cliente != None and venta_e_cliente != "":
            queryset = queryset.filter(detalle_venta__venta_e_cliente__cli_clave=venta_e_cliente)
        if venta_e_status != None and venta_e_status != "":
            queryset = queryset.filter(detalle_venta__venta_e_status=venta_e_status)
        if venta_e_tipo != None and venta_e_tipo != "":
            queryset = queryset.filter(detalle_venta__venta_e_tipo=venta_e_tipo)
        if venta_e_creado != None and venta_e_creado != "":
            queryset = queryset.filter(detalle_venta__venta_e_creado=venta_e_creado)
        
        queryset = queryset.values(
                        'detalle_producto_id__producto_marca',
                        'detalle_producto_id__producto_marca__marca_nombre',
                    ).annotate(total=Sum(F('detalle_cantidad')*F('detalle_precio'), output_field=FloatField()),).order_by('detalle_producto_id__producto_marca')
       

       

        ws['A1'] = 'TOTAL VENTAS EXPO POR MARCAS DE {} AL {}'.format(venta_e_fecha_pedido_init, venta_e_fecha_pedido_end)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:I1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Marca'
        ws['B2'] = 'Subtotal'
        # ws['C2'] = 'Actualizado'
        # ws['D2'] = 'Cliente'
        # ws['E2'] = 'Creador'
        # ws['F2'] = 'Tipo'
        # ws['G2'] = 'Total'
        # ws['H2'] = 'Status'
        cont = 3
  

        for item in queryset:
            ws.cell(row=cont, column=1).value = item['detalle_producto_id__producto_marca__marca_nombre']
            ws.cell(row=cont, column=2).value = item['total']
           
            # ws.cell(row=cont, column=8).fill = PatternFill(bgColor="00ff7f", fill_type = "solid")
            ws.cell(row=cont, column=2).number_format = '#,##0'
            cont += 1

        ws["A"+str(cont)] = "TOTAL"
        ws["B"+str(cont)] = "=SUM(B3:B"+str(cont-1)+")"
        ws["B"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'VENTAS_EXPO_MARCAS.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class TotalVentasItemsMarcasDate(TemplateView):
    def get(self, request, *args, **kwargs):

        wb = Workbook()
        ws = wb.active
        queryset = Detalle_venta.objects.all().order_by('detalle_venta')

        marca_id = self.request.GET.get('marca_id')
        fecha_init = self.request.GET.get('fecha_init')
        fecha_end = self.request.GET.get('fecha_end')
        
        marca_nom = Marca.objects.get(marca_id_marca=marca_id)
        

        if fecha_init != None and fecha_init != "":
            if fecha_end != None and fecha_end != "":
                queryset = queryset.filter(detalle_venta__venta_e_fecha_pedido__range=[fecha_init, fecha_end])
        if marca_id != None and marca_id != "":
            queryset = queryset.filter(detalle_producto_id__producto_marca=marca_id)
       
        
    

        ws['A1'] = 'TOTAL DE PORDUCTOS EXPO POR MARCAS DE {} AL {} "{}" '.format(localize(datetime.strptime(fecha_init, '%Y-%m-%d').date()), localize(datetime.strptime(fecha_end, '%Y-%m-%d').date()), marca_nom.marca_nombre)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:I1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'N° Vent'
        ws['B2'] = 'Producto'
        ws['C2'] = 'Descripción'
        ws['D2'] = 'Cantidad'
        ws['E2'] = 'Precio'
        ws['F2'] = 'Subtotal'
     
        cont = 3
  

        for item in queryset:
            ws.cell(row=cont, column=1).value = str(item.detalle_venta)
            ws.cell(row=cont, column=2).value = str(item.detalle_producto_id)
            ws.cell(row=cont, column=3).value = item.detalle_producto_id.producto_descripcion
            ws.cell(row=cont, column=4).value = item.detalle_cantidad
            ws.cell(row=cont, column=5).value = item.detalle_precio
            ws.cell(row=cont, column=6).value = item.subtotal()
           
            
            ws.cell(row=cont, column=6).number_format = '#,##0'
            cont += 1

        ws["E"+str(cont)] = "TOTAL"
        ws["F"+str(cont)] = "=SUM(F3:F"+str(cont-1)+")"
        ws["F"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        nombre_archivo = 'VENTAS_PRODUCTOS_EXPO_MARCAS.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response