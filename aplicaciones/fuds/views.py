from django.shortcuts import render
from django.views.generic import ListView, DeleteView, UpdateView, CreateView, View, TemplateView
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from aplicaciones.fuds.models import Fud,Motivo,Tramite,Conformidad,Vendedores,Zona,PartidasFud, DEVOLUCION, ESTADOS
from django.contrib import messages
from datetime import datetime, timedelta
from aplicaciones.fuds.forms import FudForm,MotivoForm,ConformidadForm,TramiteForm, FudFormEdit,FudFormEdit2,ZonaForm,VendedorForm,PartidaFudForm,Clientes,ClientForm,ClientEditForm
from django.db.models import Sum,F
from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects
from aplicaciones.pedidos.models import Producto
from django.db.models import ProtectedError, Q, F, Sum, FloatField


# pylint: disable = E1101
class MotivoCreate(CreateView):
    model= Motivo
    form_class = MotivoForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarMotivo")

    @method_decorator(permission_required('fuds.add_motivo',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MotivoCreate, self).dispatch(*args, **kwargs)

class MotivoUpdate(UpdateView):
    model= Motivo
    form_class = MotivoForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarMotivo")

    @method_decorator(permission_required('fuds.change_motivo',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MotivoUpdate, self).dispatch(*args, **kwargs)

class MotivoList(ListView):
    model= Motivo
    template_name='fuds/ViewMotivo.html'

    @method_decorator(permission_required('fuds.view_motivo',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MotivoList, self).dispatch(*args, **kwargs)

class MotivoDelete(DeleteView):
    model= Motivo
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarMotivo")

    @method_decorator(permission_required('fuds.delete_motivo',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(MotivoDelete, self).dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

class ClientCreate(CreateView):
    model= Clientes
    form_class = ClientForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarClientes")

    @method_decorator(permission_required('fuds.add_clientes',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ClientCreate, self).dispatch(*args, **kwargs)
    

class ClientList(ListView):
    model= Clientes
    template_name='fuds/ViewClient.html'

    @method_decorator(permission_required('fuds.view_clientes',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ClientList, self).dispatch(*args, **kwargs)

class ClientUpdate(UpdateView):
    model= Clientes
    form_class = ClientEditForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarClientes")

    @method_decorator(permission_required('fuds.change_clientes',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ClientUpdate, self).dispatch(*args, **kwargs)

class ClientDelete(DeleteView):
    model= Clientes
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarClientes")

    @method_decorator(permission_required('fuds.delete_clientes',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ClientDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)


        

class ConformidadCreate(CreateView):
    model= Conformidad
    form_class = ConformidadForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarConformidad")

    @method_decorator(permission_required('fuds.add_conformidad',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ConformidadCreate, self).dispatch(*args, **kwargs)

class ConformidadUpdate(UpdateView):
    model= Conformidad
    form_class = ConformidadForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarConformidad")

    @method_decorator(permission_required('fuds.change_conformidad',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ConformidadUpdate, self).dispatch(*args, **kwargs)

class ConformidadList(ListView):
    model= Conformidad
    template_name='fuds/ViewConformidad.html'

    @method_decorator(permission_required('fuds.view_conformidad',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ConformidadList, self).dispatch(*args, **kwargs)

class ConformidadDelete(DeleteView):
    model= Conformidad
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarConformidad")

    @method_decorator(permission_required('fuds.delete_conformidad',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ConformidadDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)



# CLASES DE FUD

class FudCreate(CreateView):
    model = Fud
    template_name = 'fuds/fud/create.html'
    form_class = FudForm
    success_url = reverse_lazy('fuds:fud_list')

    @method_decorator(permission_required('fuds.add_fud',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(FudCreate, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        context['tituloBrea'] = 'Crear Fud'
        return context
    def form_valid(self, form):
        form.instance.creado_por=self.request.user
        return super().form_valid(form)

class FudList(ListView): 
    model = Fud
    paginate_by = 50
    template_name= 'fuds/fud/list.html' 

    @method_decorator(permission_required('fuds.view_fud',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(FudList, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user 
        context['vendedor'] = Vendedores.objects.all().order_by('Vend_nombre')
        context['tip_accion'] = DEVOLUCION
        context['estados'] = ESTADOS
        context['conformidad'] = Conformidad.objects.all().order_by('conformidad_descripcion')
        context['motivo'] = Motivo.objects.all().order_by('motivo_descripcion')
        urls_formateada = self.request.GET.copy()
        if 'page' in urls_formateada:
            del urls_formateada['page']
        context['urls_formateada'] = urls_formateada 
        return context
    
    def get_queryset(self):
        queryset = super(FudList, self).get_queryset()
        

        fecha_captura_ini= self.request.GET.get("fecha_captura_ini")
        fecha_captura_end= self.request.GET.get("fecha_captura_end")
        vendedors= self.request.GET.get("vendedors")
        tip_aacion= self.request.GET.get("tip_aacion")
        conformidad= self.request.GET.get("conformidad")
        motivo= self.request.GET.get("motivo")
        status= self.request.GET.get("status") 
        Buscar= self.request.GET.get("Buscar") 

       


        
        if fecha_captura_ini != None and fecha_captura_ini != "":
            if fecha_captura_end != None and fecha_captura_end != "":
                queryset = queryset.filter(fecha_creacion__range=[fecha_captura_ini, fecha_captura_end])

        if vendedors != None and vendedors != "":
            queryset = queryset.filter(VendedorCliente=vendedors)

        if tip_aacion != None and tip_aacion != "":
            queryset = queryset.filter(devolucion=tip_aacion)

        if conformidad != None and conformidad != "":
            queryset = queryset.filter(Motivo__motivo_idconformidad=conformidad)

        if motivo != None and motivo != "":
            queryset = queryset.filter(Motivo=motivo)

        if status != None and status != "":
            queryset = queryset.filter(EstadoFud=status)

        if Buscar != None and Buscar != "":
            queryset = queryset.filter(Folio=Buscar)
    

            
        return queryset 



class dowload_xls_fuds(TemplateView): 
    def get(self, request , *args, **kwargs):
        from openpyxl.styles import Font, Fill, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from decimal import Decimal
        from django.utils.formats import localize
        wb = Workbook()
        ws=wb.active


        fecha_captura_ini= self.request.GET.get("fecha_captura_ini")
        fecha_captura_end= self.request.GET.get("fecha_captura_end")
        vendedors= self.request.GET.get("vendedors")
        tip_aacion= self.request.GET.get("tip_aacion")
        conformidad= self.request.GET.get("conformidad")
        motivo= self.request.GET.get("motivo")
        status= self.request.GET.get("status")
        Buscar= self.request.GET.get("Buscar") 

        queryset = Fud.objects.all()

        
        if fecha_captura_ini != None and fecha_captura_ini != "":
            if fecha_captura_end != None and fecha_captura_end != "": 
                queryset = queryset.filter(fecha_creacion__range=[fecha_captura_ini, fecha_captura_end])

        if vendedors != None and vendedors != "":
            queryset = queryset.filter(VendedorCliente=vendedors)

        if tip_aacion != None and tip_aacion != "":
            queryset = queryset.filter(devolucion=tip_aacion)

        if conformidad != None and conformidad != "":
            queryset = queryset.filter(Motivo__motivo_idconformidad=conformidad)

        if motivo != None and motivo != "":
            queryset = queryset.filter(Motivo=motivo)

        if status != None and status != "":
            queryset = queryset.filter(EstadoFud=status)

        if Buscar != None and Buscar != "":
            queryset = queryset.filter(Folio=Buscar)

        ws['A1'] = "Reporte Fud"
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:N1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = '###'
        ws['B2'] = 'Fecha Factura'
        ws['C2'] = 'Cliente'
        ws['D2'] = 'Zona'
        ws['E2'] = 'Vendedor'
        ws['F2'] = 'Conformidad'
        ws['G2'] = 'Motivo'
        ws['H2'] = 'Tramite'
        ws['I2'] = 'Devolucion'
        ws['J2'] = 'Responsable'
        ws['K2'] = 'Observaciones'
        ws['L2'] = 'Creacion'
        ws['M2'] = 'Creado'
        ws['N2'] = 'Estatus'
        cont = 3
        

        for cto in queryset:
            ws.cell(row=cont, column=1).value = cto.Folio
            ws.cell(row=cont, column=2).value = cto.FechaFactura
            ws.cell(row=cont, column=3).value = str(cto.NumeroCliente)
            ws.cell(row=cont, column=4).value = cto.VendedorCliente.Vend_Zona.Zona_nombre if cto.VendedorCliente != None else "N/a"
            ws.cell(row=cont, column=5).value = str(cto.VendedorCliente)
            ws.cell(row=cont, column=6).value = cto.Motivo.motivo_idconformidad.conformidad_descripcion
            ws.cell(row=cont, column=7).value = str(cto.Motivo)
            ws.cell(row=cont, column=8).value = str(cto.tramite)
            ws.cell(row=cont, column=9).value = cto.get_devolucion_display()
            ws.cell(row=cont, column=10).value = cto.responsable
            ws.cell(row=cont, column=11).value = cto.observaciones
            ws.cell(row=cont, column=12).value = localize(cto.fecha_creacion)
            ws.cell(row=cont, column=13).value = str(cto.creado_por)
            ws.cell(row=cont, column=14).value = cto.get_EstadoFud_display()
            
            cont += 1

     

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1


        nombre_archivo='report_fud_filters.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

    # def post(self, request, *args, **kwargs):
    #     resultado = request.POST.get("txt_search")
    #     clientes= Clientes.objects.filter(Q(Client_numero__icontains = resultado) | Q(Client_Nombre= resultado) )
    #     clientes= Clientes.objects.filter(Q(Client_numero__exact = resultado) | Q(Client_Nombre= resultado) )
    #     String = ""
    #     String2 = ""
    #     for pd in clientes:
    #         String = pd.Client_numero,
    #         String2 = pd.Client_Nombre,

    #     data = {
    #             'numerodecliente': String,
    #             'nombredecliente': String2,
    #         }
    #     return JsonResponse(data)
        


class FudUpdate(UpdateView):
    model = Fud 
    template_name = 'fuds/fud/createUpdate.html'
    form_class = FudFormEdit2
    success_url = reverse_lazy('fuds:fud_list')

    @method_decorator(permission_required('fuds.change_fud',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(FudUpdate, self).dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        identificador=self.request.GET.get('q')
        if identificador != None :
            Fud.objects.filter(Folio=identificador).update(EstadoFud=2)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        context['resultados'] = PartidasFud.objects.filter(Partida_fud = self.kwargs.get('pk'))
        context['fecha_factura']=Fud.objects.get(Folio = self.kwargs.get('pk'))
        descuento= context['fecha_factura'].Descuento
        context['total_partidas'] = PartidasFud.objects.filter(Partida_fud = self.kwargs.get('pk')).aggregate(total=Sum( F('Partida_Cantidad') * F('Partida_Precio'), output_field=FloatField() ))['total']
        context['total_descuento'] = PartidasFud.objects.filter(Partida_fud = self.kwargs.get('pk')).aggregate(total=Sum( F('Partida_Cantidad') * F('Partida_Precio'), output_field=FloatField() )*(descuento/100))['total']
        context['fecha_hoy']=datetime.now();

        if context['total_descuento'] == None :
            context['total_descuento'] = 0
            context['total_iva'] = 0
            context['total_total'] = 0
        else:
            context['total_descuento'] = round(context['total_descuento'],2)
            context['total_iva'] =(context['total_partidas']-context['total_descuento'])*0.16
            total_total = context['total_partidas']-context['total_descuento']+context['total_iva']
            context['total_iva'] = round(context['total_iva'],2)
            context['total_total'] = round(total_total,2)
        if context['total_partidas'] == None :
            context['total_partidas']=0
        else:
            context['total_partidas']=round(context['total_partidas'],2)
        return context

# class FudEnviar(TemplateView): 
#     template_name = 'fuds/fud/createUpdate.html'
#     success_url = reverse_lazy('fuds:fud_list')
#     def get(self, request, *args, **kwargs):
#         identificador=self.request.GET.get('q')
#         Fud.objects.filter(Folio=identificador).update(EstadoFud=2)
#         context = self.get_context_data(**kwargs)
#         return self.render_to_response(context)


class FudDelete(DeleteView): 
    model= Fud
    template_name='fuds/DeleteMotivo.html'
    success_url = reverse_lazy('fuds:fud_list')

    @method_decorator(permission_required('fuds.delete_fud',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(FudDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context

class TramiteCreate(CreateView):
    model= Tramite
    form_class = TramiteForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarTramites")

    @method_decorator(permission_required('fuds.add_tramite',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(TramiteCreate, self).dispatch(*args, **kwargs)

class TramiteUpdate(UpdateView):
    model= Tramite
    form_class = TramiteForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarTramites")

    @method_decorator(permission_required('fuds.change_tramite',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(TramiteUpdate, self).dispatch(*args, **kwargs)

class TramiteList(ListView):
    model= Tramite
    template_name='fuds/ViewTramite.html'

    @method_decorator(permission_required('fuds.view_tramite',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(TramiteList, self).dispatch(*args, **kwargs)

class TramiteDelete(DeleteView):
    model= Tramite
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarTramites")

    @method_decorator(permission_required('fuds.delete_tramite',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(TramiteDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context
    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)


class ZonaCreate(CreateView):
    model= Zona
    form_class = ZonaForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarZona")

    @method_decorator(permission_required('fuds.add_zona',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ZonaCreate, self).dispatch(*args, **kwargs)

class ZonaList(ListView):
    model= Zona
    template_name='fuds/ViewZona.html'

    @method_decorator(permission_required('fuds.view_zona',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ZonaList, self).dispatch(*args, **kwargs)

class ZonaUpdate(UpdateView):
    model= Zona
    form_class = ZonaForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarZona")

    @method_decorator(permission_required('fuds.change_zona',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ZonaUpdate, self).dispatch(*args, **kwargs)

class ZonaDelete(DeleteView):
    model= Zona
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarZona")

    @method_decorator(permission_required('fuds.delete_zona',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(ZonaDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context
    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)

class VendedorCreate(CreateView):
    model= Vendedores
    form_class = VendedorForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarVendedor")

    @method_decorator(permission_required('fuds.add_vendedores',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(VendedorCreate, self).dispatch(*args, **kwargs)


class VendedorList(ListView):
    model= Vendedores
    template_name='fuds/ViewVendedores.html'

    @method_decorator(permission_required('fuds.view_vendedores',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(VendedorList, self).dispatch(*args, **kwargs)

class VendedorUpdate(UpdateView):
    model= Vendedores
    form_class = VendedorForm
    template_name='fuds/CreateMotivo.html'
    success_url=reverse_lazy("fuds:ListarVendedor")

    @method_decorator(permission_required('fuds.change_vendedores',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(VendedorUpdate, self).dispatch(*args, **kwargs)

class VendedorDelete(DeleteView):
    model= Vendedores
    template_name='fuds/DeleteMotivo.html'
    success_url=reverse_lazy("fuds:ListarVendedor")

    @method_decorator(permission_required('fuds.delete_vendedores',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(VendedorDelete, self).dispatch(*args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context
    def post(self, request, *args, **kwargs):
        try:
            return self.delete(request, *args, **kwargs)
        except ProtectedError:
            contex = {
        'proveedores': 'proveedor'
                        }
        return render(request, 'pagoproveedor/protecteError.html', contex)



class PartidaCreate(CreateView):
    model= PartidasFud
    form_class = PartidaFudForm
    template_name='fuds/CreatePartida.html'
    success_url=reverse_lazy("fuds:ListarVendedor")

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the books
        context['usuario'] = self.request.user
        context['resultados'] =productoFiltrado= PartidasFud.objects.filter(Partida_fud = self.kwargs.get('pk'))

        return context

    @method_decorator(permission_required('fuds.add_vendedores',reverse_lazy('inicio:need_permisos')))
    def dispatch(self, *args, **kwargs):
                return super(PartidaCreate, self).dispatch(*args, **kwargs)
    def post(self, request, *args, **kwargs):
        producto=Producto.objects.get(producto_codigo=request.POST.get('Partida_nombre'))
        pf=PartidasFud(
                Partida_nombre=producto,
                Partida_fud_id=request.POST.get('Partida_fud'),
                Partida_Precio=request.POST.get('Partida_Precio'),
                Partida_Cantidad=request.POST.get('Partida_Cantidad'),
                )
        pf.save()
        respuesta="""
        <div class="alert alert-success alert-dismissible fade show" role="alert">
            {Partida_nombre} agredado al FUD!
            <button type="button" class="close" data-dismiss="alert" aria-label="Close">
                <span aria-hidden="true">&times;</span>
            </button>
        </div>
        """.format(
            Partida_nombre = request.POST.get('Partida_nombre'),
        )
        productoFiltrado= PartidasFud.objects.filter(Partida_fud = request.POST.get('Partida_fud'))
        String2= ""
        for pd2 in productoFiltrado:
            String2 += """
            <tr>
                <th scope="row">{Partida_nombre}</th>
                <th scope="row">{Partida_fud}</th>
                <th scope="row">{Partida_Precio}</th>
                <th scope="row">{Partida_Cantidad}</th>
                
            </tr>
            """.format(
                Partida_nombre = pd2.Partida_nombre,
                Partida_fud = pd2.Partida_fud,
                Partida_Precio = pd2.Partida_Precio,
                Partida_Cantidad = pd2.Partida_Cantidad,
            )

        return JsonResponse({'cp':String2,'rsp':respuesta})


class PartidaView(View):
    
    # template_name='fuds/ViewVendedores.html'
    def post(self, request, *args, **kwargs):
        resultado = request.POST.get("txt_search")
        idfud = request.POST.get("txt_idfud")
        producto= Producto.objects.filter(Q(producto_descripcion__icontains = resultado) | Q(producto_codigo= resultado) )
        productoFiltrado= PartidasFud.objects.filter(Partida_fud = idfud )
        String = """
         <thead class="thead-dark">
                <tr>
                <th scope="col">Partida</th>
                <th scope="col">Descripción</th>
                <th scope="col">Precio</th>
                <th scope="col">Cantidad</th>
                <th scope="col">Acciones</th>
                </tr>
            </thead>"""
        for pd in producto:
            String += """
            <form action="asdasdasd.com" method="POST" >
            <tr>
                <th scope="row">{Partida}</th>
                <th scope="row">{Descripcion} </th>
                <th scope="row"><input type="number" step="any" class="form-control" id="Partida_Precio{Partida}"></th>
                <th scope="row"><input type="number" class="form-control" id="Partida_Cantidad{Partida}"></th>
                <th> <input type="submit" class="btn btn-info" onclick="guardar_partida('{Partida}',{idfud})" value="Agregar a fud" placeholder="Busqueda de producto"> </th>
            </tr>
            </form>
            """.format(
                Partida = pd.producto_codigo,
                Descripcion = pd.producto_descripcion,
                idfud = idfud,
            )

        String2= ''
        for pd2 in productoFiltrado:
            String2 += """
            <tr>
                <th scope="row">{Partida_nombre}</th>
                <th scope="row">{Partida_fud}</th>
                <th scope="row">{Partida_Precio}</th>
                <th scope="row">{Partida_Cantidad}</th>
            </tr>
            """.format(
                Partida_nombre = pd2.Partida_nombre,
                Partida_fud = pd2.Partida_fud,
                Partida_Precio = pd2.Partida_Precio,
                Partida_Cantidad = pd2.Partida_Cantidad,
            )

        data = {
                'cuerpoT': String,
                'cuerpoF': String2,
            }
        return JsonResponse(data)

class ClienteVerCaptura(View):
    # template_name='fuds/ViewVendedores.html'
    def post(self, request, *args, **kwargs):
        resultado = request.POST.get("txt_search")
        # clientes= Clientes.objects.filter(Q(Client_numero__icontains = resultado) | Q(Client_Nombre= resultado) )
        clientes= Clientes.objects.filter(Q(Client_numero__exact = resultado) | Q(Client_Nombre= resultado) )
        String = ""
        String2 = ""
        for pd in clientes:
            String = pd.Client_numero,
            String2 = pd.Client_Nombre,

        data = {
                'numerodecliente': String,
                'nombredecliente': String2,
            }
        return JsonResponse(data)