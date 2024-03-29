from django.contrib.auth.mixins import PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, ListView, DeleteView, TemplateView
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from aplicaciones.pago_proveedor.eliminaciones import get_deleted_objects
from aplicaciones.solicitud.forms import *


class TipoServicioCreate(CreateView, PermissionRequiredMixin):
    permission_required = 'solicitud.add_tiposervicio'
    model = TipoServicio
    template_name = "solicitud/tipo_servicio_create.html"
    form_class = TipoServicioForm
    success_url = reverse_lazy('solicitud:crear_tipo_servicio')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tiposervicio_list'] = TipoServicio.objects.all()
        return context


class TipoServicioEdit(UpdateView, PermissionRequiredMixin):
    permission_required = 'solicitud.change_tiposervicio'
    model = TipoServicio
    template_name = "solicitud/tipo_servicio_create.html"
    form_class = TipoServicioForm
    success_url = reverse_lazy('solicitud:crear_tipo_servicio')


class TipoServicioDelete(DeleteView, PermissionRequiredMixin):
    permission_required = 'solicitud.delete_tiposervicio'
    model = TipoServicio
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('solicitud:crear_tipo_servicio')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class ServicioList(ListView, PermissionRequiredMixin):
    permission_required = 'solicitud.view_servicio'
    model = Servicio
    template_name = "solicitud/solicitud_list.html"
    paginate_by = 50

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.order_by('s_estatus')

        s_fecha = self.request.GET.get('s_fecha')
        s_fecha2 = self.request.GET.get('s_fecha2')
        s_tipo = self.request.GET.get('s_tipo')
        s_empresa = self.request.GET.get('s_empresa')
        s_estatus = self.request.GET.get('s_estatus')

        if s_fecha and s_fecha2:
            queryset = queryset.filter(s_fecha__range=(s_fecha, s_fecha2))
        if s_tipo:
            queryset = queryset.filter(s_tipo=s_tipo)
        if s_empresa:
            queryset = queryset.filter(s_empresa=s_empresa)
        if s_estatus:
            queryset = queryset.filter(s_estatus=s_estatus)

        queryset = queryset.order_by('-s_fecha')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['forms_filtro'] = FiltroServicioForm(self.request.GET)
        return context

class Download_report_servicio(TemplateView):
    def get(self, request, *args, **kwargs):
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active

        queryset = Servicio.objects.all()

        s_fecha = self.request.GET.get('s_fecha')
        s_fecha2 = self.request.GET.get('s_fecha2')
        s_tipo = self.request.GET.get('s_tipo')
        s_empresa = self.request.GET.get('s_empresa')
        s_estatus = self.request.GET.get('s_estatus')

        if s_fecha and s_fecha2:
            queryset = queryset.filter(s_fecha__range=(s_fecha, s_fecha2))
        if s_tipo:
            queryset = queryset.filter(s_tipo=s_tipo)
        if s_empresa:
            queryset = queryset.filter(s_empresa=s_empresa)
        if s_estatus:
            queryset = queryset.filter(s_estatus=s_estatus)

        queryset = queryset.order_by('-s_fecha')

        ws['A1'] = 'REPORTE DE SOLICITUDES DE SERVICIO'
        st = ws['A1']
        st.font = Font(size=14, b=True, color="0A15A4")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"
        
        ws['A2'] = 'FOLIO SERVICIO'
        ws['B2'] = 'FECHA'
        ws['C2'] = 'TIPO DE SERVICIO'
        ws['D2'] = 'USUARIO ACTUALIZÓ'
        ws['E2'] = 'POR'
        ws['F2'] = 'ESTADO'
        ws['G2'] = '# SERIE'
        ws['H2'] = 'DESCRIPCIÓN SERVICIO'
        cont = 3

        for item in queryset:
            ws.cell(row=cont, column=1).value = str(item.id)
            ws.cell(row=cont, column=2).value = item.s_fecha
            ws.cell(row=cont, column=3).value = str(item.s_tipo)
            ws.cell(row=cont, column=4).value = str(item.s_user_cambio)
            ws.cell(row=cont, column=5).value = str(item.s_user)
            ws.cell(row=cont, column=6).value = item.s_estatus
            ws.cell(row=cont, column=7).value = str(item.s_serie)
            ws.cell(row=cont, column=8).value = str(item.s_reporte)
            cont += 1
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 1
    
        nombre_archivo = 'Reporte de solicitudes.xlsx'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ServicioCreate(CreateView, PermissionRequiredMixin):
    permission_required = 'solicitud.add_servicio'
    model = Servicio
    template_name = "solicitud/generic_form.html"
    form_class = ServicioForm
    success_url = reverse_lazy('solicitud:servicios')

    def form_valid(self, form):
        form.instance.s_depo = self.request.user.departamento
        form.instance.s_user = self.request.user
        form.instance.s_user_cambio = self.request.user
        self.object = form.save()
        return super().form_valid(form)


class ServicioValidarView(UpdateView, PermissionRequiredMixin):
    permission_required = 'solicitud.validar_servicio'
    model = Servicio
    template_name = "solicitud/generic_form.html"
    form_class = ServicioValidarForm
    success_url = reverse_lazy('solicitud:servicios')

    def form_valid(self, form):
        form.instance.s_estatus = 2
        form.instance.s_user_cambio = self.request.user
        self.object = form.save()
        return super().form_valid(form)

class ServicioCerrarView(UpdateView, PermissionRequiredMixin):
    permission_required = 'solicitud.cerrar_servicio'
    model = Servicio
    template_name = "solicitud/generic_form.html"
    form_class = ServicioCerrarForm
    success_url = reverse_lazy('solicitud:servicios')

    def form_valid(self, form):
        form.instance.s_estatus = 4
        form.instance.s_user_cambio = self.request.user
        self.object = form.save()
        return super().form_valid(form)


class ServicioDelete(DeleteView, PermissionRequiredMixin):
    permission_required = 'solicitud.delete_servicio'
    model = Servicio
    template_name = "pedidos/delete_forever.html"
    success_url = reverse_lazy('solicitud:servicios')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class UpdateStatusServicioView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = ()

    def post(self, request, *args, **kwargs):
        servicio_id = request.data["id"]
        user = request.data["changue"]
        resp = Servicio.objects.filter(id=servicio_id).update(s_estatus=3, s_user_cambio=user)
        data = {
            'respuesta': resp,
        }
        return Response(data, status=status.HTTP_202_ACCEPTED)
