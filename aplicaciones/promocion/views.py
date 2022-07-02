from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView, ListView
from aplicaciones.promocion.models import *
from django.contrib import messages
from django.shortcuts import redirect
from datetime import datetime
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Avg, Sum, F, FloatField, Count, Q, Case, When, Max
# Create your views here.
class ProductosInsertPromo(TemplateView):
    template_name = "promocion/update_bulk.html"

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_superuser == False:
            redirect('promocion:update_masive_promo')

        return super(ProductosInsertPromo, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        return context

    def post(self, request, *args, **kwargs):
        import openpyxl
        file_producto_xls = request.FILES['file_producto_xls']
        wb = openpyxl.load_workbook(file_producto_xls)
        # getting a particular sheet by name out of many sheets
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]
        # print(worksheet)

        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        excel_data.pop(0)
        bulk_list = list()
        for item in excel_data:
            if item[14] == 'F':
                bulk_list.append(Promocion(
                        venta=item[0],
                        no_cliente=item[1],
                        cod_prod=item[3],
                        descripcion=item[4],
                        importeNeto=item[8],
                        fecha_fac=datetime.strptime(item[9], '%Y-%m-%d %H:%M:%S').date(),
                        fac=item[10],
                        proveedor=item[12],
                    ))
                # esp.save()
                # Promocion.objects.filter(producto_codigo=item[0]).update(
                #     producto_precio=item[1])
        Promocion.objects.bulk_create(bulk_list,5000)
        messages.success(request, 'Registros realizados correctamente.')
        return redirect('promocion:update_masive_promo')

class PromoList(LoginRequiredMixin,ListView):
    login_url = reverse_lazy('inicio')
    redirect_field_name = 'redirect_to'
    template_name = 'web/V2/promo_lista.html'
    model = Promocion
    paginate_by = 24
    ordering = ['-venta']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['puntos']=Promocion.objects.filter(no_cliente=self.request.user).values('proveedor').annotate(totalpuntos = Sum(F('importeNeto'),output_field=models.FloatField()), puntosR = Sum(('importeNeto'),output_field=models.FloatField())/10000 )
        # context['orden']=Promocion.objects.values('no_cliente','proveedor').annotate(totalimporte = Sum(F('importeNeto'),output_field=models.FloatField()), puntosR = Case(
        #     When(proveedor="440", then= (Sum(F('importeNeto'))/10000)*15),
        #     When(proveedor="623", then= (Sum(F('importeNeto'))/10000)*6),
        #     When(proveedor="255", then= (Sum(F('importeNeto'))/10000)*15),
        #     When(proveedor="855", then= (Sum(F('importeNeto'))/10000)*15),
        #     When(proveedor="261", then= (Sum(F('importeNeto'))/10000)*9),
        #     When(proveedor="022", then= (Sum(F('importeNeto'))/10000)*9),
        #     When(proveedor="009", then= (Sum(F('importeNeto'))/10000)*9),
        #     When(proveedor="111", then= (Sum(F('importeNeto'))/10000)*3),
        #     When(proveedor="801", then= (Sum(F('importeNeto'))/10000)*3),
        #     default=Sum(F('importeNeto')),
        # ))
        context['orden']=Promocion.objects.all()
        
        context['mejores']=Promocion.objects.values('no_cliente').annotate(totalimporte = Sum(F('importeNeto'),output_field=models.FloatField()) ).order_by('-totalimporte')[:10]
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(no_cliente=self.request.user).values('fac').annotate(totalVenta = Sum(F('importeNeto'),output_field=models.FloatField()))
        # queryset = queryset.filter(no_cliente=self.request.user)
        return queryset

class PromoListDetalle(LoginRequiredMixin,ListView):
    login_url = reverse_lazy('inicio')
    redirect_field_name = 'redirect_to'
    template_name = 'web/V2/promo_detalle.html'
    model = Promocion
    paginate_by = 24
    ordering = ['-venta']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(fac=self.kwargs['nombre'])
        return queryset